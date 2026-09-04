"""
Report-7 — Performance in Terms of Vessels Handled - All Vessels
Flask Blueprint version. Reads from mis_vessel_master (historical) and falls back
to live operational pipeline (lueu_parcel_log + ldud_header + vcn_header) for post-June open periods.

Port: JAWAHARLAL NEHRU PORT AUTHORITY
Columns: Liquid Bulk | All

Section Breakdown:
  1. No. of Vessels Sailed- (Total No.)
     a) At Berth- (No.)
        (i) Overseas
        (ii) Coastal
     b) At Stream- (No.) (Inn Anchorage)
  2. Total Cargo Traffic Handled - (000 Tonnes)
     a) At Berth - (000 Tonnes)
     b) At Stream (Inn Anchorage) - (000 Tonnes)
  3. Pre Berthing Waiting Time(Hrs)-Total
     a) Pre Berthing Waiting Time(Hrs)-Port a/c
     b) Pre Berthing Waiting Time(Hrs)-Non Port a/c
  4. Working Time (Hrs)  [Formula: (first start parcel - last end parcel)]
  5. N.W Time at working berth(Hrs)-Total  [Formula: (Alongside - Castoff in hrs) - Sr No 4]
     a) N.W Time at working berth(Hrs)-Port a/c
     b) N.W Time at working berth(Hrs)-N.P a/c
  6. Time at Non working berth(Hrs)-Total
     a) Time at Non working berth(Hrs)-Port a/c
     b) Time at Non working berth(Hrs)-N.P a/c
  7. Navigation Time (Hrs)-Total
     a) Inward Movement (Hrs)  [Pilot pick up to alongside]
     b) Outward Movement (Hrs) [Pilot disembark to castoff]
  8. Shifting Time
  9. Turn Round Time (Hrs)-Total=(3+4+5+6+7+8)
     a) Turn Round Time (Hrs)-Port a/c=(3(a)+4+5(a)+6(a)+7+8)
     b) Turn Round Time (Hrs)-N.P a/c=(3(b)+5(b)+6(b))
"""

import io
import math
import traceback
from functools import wraps
from datetime import datetime, date

import pandas as pd
from flask import jsonify, request, render_template, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from database import get_db, get_cursor
from .. import bp


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]


def fy_start_year(fin_year: str) -> int:
    try:
        return int(fin_year.split("-")[0])
    except Exception:
        return datetime.now().year


def month_options_for(fin_year: str):
    start_y = fy_start_year(fin_year)
    opts = []
    for idx, mn in enumerate(MONTH_NAMES):
        yy = start_y if idx < 9 else start_y + 1
        opts.append({"idx": idx, "label": f"{mn}-{str(yy % 100).zfill(2)}"})
    return opts


def _parse_dt(s):
    if not s:
        return None
    if isinstance(s, (datetime, date)):
        return s
    try:
        return datetime.strptime(str(s).replace('T', ' ')[:16], '%Y-%m-%d %H:%M')
    except Exception:
        try:
            return datetime.strptime(str(s)[:10], '%Y-%m-%d')
        except Exception:
            return None


def _load_live_pipeline_rows():
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT ld.id AS ldud_id,
                   ld.cast_off_datetime,
                   ld.alongside_datetime,
                   ld.anchored_datetime,
                   ld.nor_tendered,
                   ld.discharge_commenced,
                   ld.discharge_completed,
                   ld.pilot_pickup_time,
                   ld.pilot_disembarked,
                   h.id AS vcn_id,
                   h.vessel_name,
                   h.berth_name,
                   h.cargo_type,
                   h.operation_type,
                   SUM(l.quantity) AS quantity,
                   MIN(po.start_dt) AS first_parcel_start,
                   MAX(po.end_dt) AS last_parcel_end
            FROM lueu_parcel_log l
            JOIN ldud_parcel_ops po ON po.id = l.parcel_op_id
            JOIN ldud_header ld ON ld.id = po.ldud_id
            JOIN vcn_header h ON h.id = ld.vcn_id
            WHERE l.is_deleted IS NOT TRUE
              AND COALESCE(l.is_shortclose, FALSE) = FALSE
            GROUP BY ld.id,
                     ld.cast_off_datetime,
                     ld.alongside_datetime,
                     ld.anchored_datetime,
                     ld.nor_tendered,
                     ld.discharge_commenced,
                     ld.discharge_completed,
                     ld.pilot_pickup_time,
                     ld.pilot_disembarked,
                     h.id,
                     h.vessel_name,
                     h.berth_name,
                     h.cargo_type,
                     h.operation_type
        """)
        live_raw = cur.fetchall()
    finally:
        conn.close()

    live_rows = []
    for r in live_raw:
        dt = _parse_dt(r['cast_off_datetime']) or _parse_dt(r['alongside_datetime'])
        if dt:
            fy_start = dt.year if dt.month >= 4 else dt.year - 1
            fy = f"{fy_start}-{(fy_start + 1) % 100:02d}"
            idx = dt.month - 4 if dt.month >= 4 else dt.month + 8
            mn_label = f"{MONTH_NAMES[idx]}-{str(dt.year)[-2:]}"

            anchored = _parse_dt(r['anchored_datetime'])
            along = _parse_dt(r['alongside_datetime'])
            cast = _parse_dt(r['cast_off_datetime'])
            p_pick = _parse_dt(r['pilot_pickup_time'])
            p_dis = _parse_dt(r['pilot_disembarked'])

            # Working Time for live pipeline:
            # First parcel operation start -> Last parcel operation end
            first_parcel_start = _parse_dt(r.get('first_parcel_start'))
            last_parcel_end = _parse_dt(r.get('last_parcel_end'))

            wt_hrs = (
                (last_parcel_end - first_parcel_start).total_seconds() / 3600.0
                if first_parcel_start
                and last_parcel_end
                and last_parcel_end > first_parcel_start
                else 0.0
            )

            # 3. Pre-Berthing Waiting Time
            # Approved formula:
            # Pre-Berthing Waiting Time = Alongside Date/Time - Anchorage Date/Time
            pbw_hrs = (
                (along - anchored).total_seconds() / 3600.0
                if anchored and along and along > anchored
                else 0.0
            )

            # 5. Total Berth Stay
            # Approved formula:
            # Total Berth Stay = Cast-off / Sail Date/Time - Alongside Date/Time
            sab_hrs = (
                (cast - along).total_seconds() / 3600.0
                if cast and along and cast > along
                else 0.0
            )

            # 7. Navigation Time
            # Approved formula:
            # Navigation Time = Alongside Date/Time - Pilot Pick-up Date/Time
            inward_hrs = (
                (along - p_pick).total_seconds() / 3600.0
                if along and p_pick and along > p_pick
                else 0.0
            )

            # Outward Movement = Cast-off - Pilot Disembarked
            outward_hrs = (
                (cast - p_dis).total_seconds() / 3600.0
                if cast and p_dis and cast > p_dis
                else 0.0
            )

            live_rows.append({
                'fin_year': fy,
                'month': mn_label,
                'berth_no': r['berth_name'],
                'cargo': r['cargo_type'],
                'category': 'LIQUID',
                'new_cat': 'LIQUID',
                'category1': 'LIQUID',
                'unloading_terminal': r['berth_name'],
                'quantity': float(r['quantity'] or 0),
                'pre_berthing_waiting': pbw_hrs / 24.0,
                'waiting_port': pbw_hrs / 24.0,
                'waiting_non_port': 0.0,
                'stay_at_berth': sab_hrs / 24.0,
                'working_time': wt_hrs / 24.0,
                'inward_movement': inward_hrs / 24.0,
                'outward_movement': outward_hrs / 24.0,
                'overseas_coastal': 'Overseas',
                'vcn_no': str(r['vcn_id']),
                'ops_commenced': first_parcel_start,
                'cargo_completion': last_parcel_end,
                'alongside': along,
                'cast_off': cast,
                'pilot_pickup': p_pick,
                'pilot_disembarked': p_dis
            })
    return live_rows


def calculate_section_metrics(rlist):
    if not rlist:
        return {
            "vsl_sailed_total": 0,
            "vsl_berth_total": 0,
            "vsl_overseas": 0,
            "vsl_coastal": 0,
            "vsl_stream_total": 0,
            "traffic_total": 0.0,
            "traffic_berth": 0.0,
            "traffic_stream": 0.0,
            "pbw_total": 0.0,
            "pbw_port": 0.0,
            "pbw_non_port": 0.0,
            "working_time": 0.0,
            "nw_working_total": 0.0,
            "nw_working_port": 0.0,
            "nw_working_non_port": 0.0,
            "nw_non_working_total": 0.0,
            "nw_non_working_port": 0.0,
            "nw_non_working_non_port": 0.0,
            "nav_total": 0.0,
            "inward_movement": 0.0,
            "outward_movement": 0.0,
            "shifting_time": 0.0,
            "trt_total": 0.0,
            "trt_port": 0.0,
            "trt_non_port": 0.0
        }

    # 1. Vessels Sailed
    vsl_sailed_total = len(rlist)
    vsl_stream_total = sum(1 for r in rlist if 'ANCH' in str(r.get('berth_no') or '').upper() or 'STREAM' in str(r.get('berth_no') or '').upper())
    vsl_berth_total = vsl_sailed_total - vsl_stream_total
    vsl_coastal = sum(1 for r in rlist if str(r.get('overseas_coastal') or '').strip().lower() == 'coastal' and 'ANCH' not in str(r.get('berth_no') or '').upper())
    vsl_overseas = vsl_berth_total - vsl_coastal

    # 2. Traffic Handled (000 Tonnes)
    traffic_total = sum(float(r.get('quantity') or 0) for r in rlist) / 1000.0
    traffic_stream = sum(float(r.get('quantity') or 0) for r in rlist if 'ANCH' in str(r.get('berth_no') or '').upper() or 'STREAM' in str(r.get('berth_no') or '').upper()) / 1000.0
    traffic_berth = traffic_total - traffic_stream

    # 3. Pre Berthing Waiting Time (Hrs)
    pbw_non_port = sum(float(r.get('waiting_non_port') or 0) for r in rlist) * 24.0
    pbw_port = sum(float(r.get('waiting_port') or 0) for r in rlist) * 24.0
    pbw_total_col = sum(float(r.get('pre_berthing_waiting') or 0) for r in rlist) * 24.0

    if pbw_port == 0 and pbw_non_port == 0:
        pbw_total = round(pbw_total_col, 2)
        pbw_port = pbw_total
    else:
        pbw_total = round(pbw_total_col if pbw_total_col > (pbw_port + pbw_non_port) else (pbw_port + pbw_non_port), 2)
        pbw_port = round(pbw_total - pbw_non_port, 2)
        pbw_non_port = round(pbw_non_port, 2)

    # 4. Working Time (in Days and Hrs)
    wt_days_sum = 0.0
    wt_hrs_sum = 0.0
    for r in rlist:
        ops_start = _parse_dt(r.get('ops_commenced'))
        ops_end = _parse_dt(r.get('cargo_completion'))
        if ops_start and ops_end and ops_end > ops_start:
            wt_hrs_sum += (ops_end - ops_start).total_seconds() / 3600.0
            wt_days_sum += (ops_end - ops_start).total_seconds() / 86400.0
        else:
            wt_hrs_sum += float(r.get('working_time') or 0) * 24.0
            wt_days_sum += float(r.get('working_time') or 0)

    working_time = round(wt_days_sum, 2)

    # 5. Stay at berth in hours (using stay_at_berth column if present, else Cast Off - Alongside timestamp)
    sab_hrs_sum = 0.0
    for r in rlist:
        if r.get('stay_at_berth') is not None:
            sab_hrs_sum += float(r.get('stay_at_berth') or 0) * 24.0
        else:
            along = _parse_dt(r.get('alongside'))
            cast = _parse_dt(r.get('cast_off'))
            if along and cast and cast > along:
                sab_hrs_sum += (cast - along).total_seconds() / 3600.0

    # N.W Time at working berth (Hrs) = Stay at berth (in Hrs) - Working Time (in Hrs)
    nw_working_total = round(sab_hrs_sum - wt_hrs_sum, 2)
    if nw_working_total < 0:
        nw_working_total = 0.0
    nw_working_port = nw_working_total
    nw_working_non_port = 0.0

    # 6. Time at Non working berth
    nw_non_working_total = 0.0
    nw_non_working_port = 0.0
    nw_non_working_non_port = 0.0

    # 7. Navigation Time (Hrs)
    in_ts_sum = 0.0
    out_ts_sum = 0.0
    for r in rlist:
        along = _parse_dt(r.get('alongside'))
        cast = _parse_dt(r.get('cast_off'))
        p_pick = _parse_dt(r.get('pilot_pickup'))
        p_dis = _parse_dt(r.get('pilot_disembarked'))

        if p_pick and along and along > p_pick:
            in_ts_sum += (along - p_pick).total_seconds() / 3600.0
        else:
            in_ts_sum += float(r.get('inward_movement') or 0) * 24.0

        if p_dis and cast and p_dis > cast:
            out_ts_sum += (p_dis - cast).total_seconds() / 3600.0
        else:
            out_ts_sum += float(r.get('outward_movement') or 0) * 24.0

    inward_movement = round(in_ts_sum, 2)
    outward_movement = round(out_ts_sum, 2)
    nav_total = round(inward_movement + outward_movement, 2)

    # 8. Shifting Time
    shifting_time = 0.0

    # 9. Turn Round Time (Hrs)
    # Total = (3 + 4 + 5 + 6)
    # Port  = (3(a) + 4 + 5(a) + 6(a) + 7 + 8)
    # N.P   = (3(b) + 5(b) + 6(b))
    trt_total = round(pbw_total + working_time + nw_working_total + nw_non_working_total, 2)
    trt_port = round(pbw_port + working_time + nw_working_port + nw_non_working_port + nav_total + shifting_time, 2)
    trt_non_port = round(pbw_non_port + nw_working_non_port + nw_non_working_non_port, 2)

    return {
        "vsl_sailed_total": vsl_sailed_total,
        "vsl_berth_total": vsl_berth_total,
        "vsl_overseas": vsl_overseas,
        "vsl_coastal": vsl_coastal,
        "vsl_stream_total": vsl_stream_total,
        "traffic_total": round(traffic_total, 3),
        "traffic_berth": round(traffic_berth, 3),
        "traffic_stream": round(traffic_stream, 3),
        "pbw_total": pbw_total,
        "pbw_port": pbw_port,
        "pbw_non_port": pbw_non_port,
        "working_time": working_time,
        "nw_working_total": nw_working_total,
        "nw_working_port": nw_working_port,
        "nw_working_non_port": nw_working_non_port,
        "nw_non_working_total": nw_non_working_total,
        "nw_non_working_port": nw_non_working_port,
        "nw_non_working_non_port": nw_non_working_non_port,
        "nav_total": nav_total,
        "inward_movement": inward_movement,
        "outward_movement": outward_movement,
        "shifting_time": shifting_time,
        "trt_total": trt_total,
        "trt_port": trt_port,
        "trt_non_port": trt_non_port
    }


def build_report7_data(fin_year: str, selected_month: str = "Jun-26"):
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT fin_year, month, berth_no, cargo, category, new_cat, category1, unloading_terminal,
                   quantity, pre_berthing_waiting, waiting_port, waiting_non_port,
                   stay_at_berth, working_time, non_working_port, inward_movement, outward_movement,
                   overseas_coastal, cast_off, vcn_no, ops_commenced, cargo_completion, alongside,
                   pilot_pickup, pilot_disembarked, nor
            FROM mis_vessel_master
            WHERE fin_year = %s
        """, [fin_year])
        mv_rows = cur.fetchall()

        cur.execute("SELECT DISTINCT fin_year FROM mis_vessel_master WHERE fin_year IS NOT NULL ORDER BY fin_year DESC")
        avail_years = [r['fin_year'] for r in cur.fetchall()]
    finally:
        conn.close()

    covered_months = set(str(r.get('month') or '').strip() for r in mv_rows if r.get('month'))

    live_pipeline_rows = _load_live_pipeline_rows()
    live_uncovered = [r for r in live_pipeline_rows if r['fin_year'] == fin_year and r['month'] not in covered_months]

    all_dataset_rows = list(mv_rows) + live_uncovered

    live_fys = set(r['fin_year'] for r in live_pipeline_rows)
    for fy in live_fys:
        if fy not in avail_years:
            avail_years.append(fy)
    avail_years.sort(reverse=True)

    m_opts = month_options_for(fin_year)
    m_labels = [opt["label"] for opt in m_opts]

    # Filter by selected_month if specified and not 'ALL'
    if selected_month and selected_month != 'ALL':
        filtered_rows = [r for r in all_dataset_rows if str(r.get('month') or '').strip() == selected_month]
    else:
        filtered_rows = all_dataset_rows

    def is_liquid(r):
        b = str(r.get('berth_no') or '').strip().upper()
        c = str(r.get('cargo') or '').strip().upper()
        cat = str(r.get('category') or '').strip().upper()
        ncat = str(r.get('new_cat') or '').strip().upper()
        term = str(r.get('unloading_terminal') or '').strip().upper()
        return ('LB-' in b) or ('LIQUID' in cat) or ('LIQUID' in ncat) or ('OIL' in c) or ('ACID' in c) or ('CHEMICAL' in cat)

    liquid_rows = [r for r in filtered_rows if is_liquid(r)]

    liquid_metrics = calculate_section_metrics(liquid_rows)
    all_metrics = calculate_section_metrics(filtered_rows)

    return {
        "fin_year": fin_year,
        "selected_month": selected_month,
        "available_years": avail_years if avail_years else [fin_year],
        "month_labels": m_labels,
        "liquid_bulk": liquid_metrics,
        "all_vessels": all_metrics
    }


def _get_current_month():
    now = datetime.now()
    idx = now.month - 4 if now.month >= 4 else now.month + 8
    return f"{MONTH_NAMES[idx]}-{str(now.year)[-2:]}"


def _get_current_fy():
    now = datetime.now()
    fy_start = now.year if now.month >= 4 else now.year - 1
    return f"{fy_start}-{(fy_start + 1) % 100:02d}"


@bp.route('/module/RP01/report7/')
@login_required
def report7_page():
    return render_template('report7/report7.html', username=session.get('username'))


@bp.route('/api/module/RP01/report7/data')
@login_required
def report7_data():
    fin_year = request.args.get('year', _get_current_fy()).strip()
    month = request.args.get('month', _get_current_month()).strip()
    try:
        data = build_report7_data(fin_year, month)
        return jsonify(data)
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/api/module/RP01/report7/export')
@login_required
def report7_export():
    fin_year = request.args.get('year', _get_current_fy()).strip()
    month = request.args.get('month', _get_current_month()).strip()
    try:
        data = build_report7_data(fin_year, month)

        wb = Workbook()
        ws = wb.active
        ws.title = "Physical Perf"
        ws.views.sheetView[0].showGridLines = True

        thin_border = Border(
            left=Side(style='thin', color='A0A0A0'),
            right=Side(style='thin', color='A0A0A0'),
            top=Side(style='thin', color='A0A0A0'),
            bottom=Side(style='thin', color='A0A0A0')
        )
        bold_font = Font(name='Calibri', size=11, bold=True)
        norm_font = Font(name='Calibri', size=10)

        title_font = Font(name='Calibri', size=14, bold=True)
        subhead_font = Font(name='Calibri', size=11, bold=True)

        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Row 1: Main Header
        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = "Performance in Terms of Vessels Handled -All Vessels"
        title_cell.font = title_font
        title_cell.alignment = left_align

        # Row 2: Sub Header
        ws.merge_cells("A2:C2")
        sub_cell = ws["A2"]
        sub_cell.value = "Port: JAWAHARLAL NEHRU PORT AUTHORITY"
        sub_cell.font = subhead_font
        sub_cell.alignment = left_align

        # Row 4: Column Headers
        ws.cell(row=4, column=1, value="Item").font = bold_font
        ws.cell(row=4, column=1).alignment = center_align
        ws.cell(row=4, column=1).border = thin_border

        ws.cell(row=4, column=2, value="Liquid\nBulk").font = bold_font
        ws.cell(row=4, column=2).alignment = center_align
        ws.cell(row=4, column=2).border = thin_border

        ws.cell(row=4, column=3, value="All").font = bold_font
        ws.cell(row=4, column=3).alignment = center_align
        ws.cell(row=4, column=3).border = thin_border

        rows_def = [
            ("1. No. of Vessels Sailed- (Total No.)", "vsl_sailed_total", "0", True, True),
            ("    a) At Berth- (No.)", "vsl_berth_total", "0", False, True),
            ("        (i) Overseas", "vsl_overseas", "0", False, True),
            ("        (ii) Coastal", "vsl_coastal", "0", False, True),
            ("    b) At Stream- (No.) (Inn Anchorage)", "vsl_stream_total", "0", False, False),
            ("2. Total Cargo Traffic Handled - (000 Tonnes)", "traffic_total", "#,##0.000", True, True),
            ("    a) At Berth - (000 Tonnes)", "traffic_berth", "#,##0.000", False, True),
            ("    b) At Stream (Inn Anchorage) - (000 Tonnes)", "traffic_stream", "#,##0.000", False, False),
            ("3. Pre Berthing Waiting Time(Hrs)-Total", "pbw_total", "#,##0.00", True, True),
            ("    a) Pre Berthing Waiting Time(Hrs)-Port a/c", "pbw_port", "#,##0.00", False, True),
            ("    b) Pre Berthing Waiting Time(Hrs)-Non Port a/c", "pbw_non_port", "#,##0.00", False, True),
            ("4. Working Time (Hrs)", "working_time", "#,##0.00", True, True),
            ("5. N.W Time at working berth(Hrs)-Total", "nw_working_total", "#,##0.00", True, True),
            ("    a) N.W Time at working berth(Hrs)-Port a/c", "nw_working_port", "#,##0.00", False, True),
            ("    b) N.W Time at working berth(Hrs)-N.P a/c", "nw_working_non_port", "#,##0.00", False, False),
            ("6. Time at Non working berth(Hrs)-Total", "nw_non_working_total", "#,##0.00", True, False),
            ("    a) Time at Non working berth(Hrs)-Port a/c", "nw_non_working_port", "#,##0.00", False, False),
            ("    b) Time at Non working berth(Hrs)-N.P a/c", "nw_non_working_non_port", "#,##0.00", False, False),
            ("7. Navigation Time (Hrs)-Total", "nav_total", "#,##0.00", True, True),
            ("    a) Inward Movement (Hrs)", "inward_movement", "#,##0.00", False, True),
            ("    b) Outward Movement (Hrs)", "outward_movement", "#,##0.00", False, True),
            ("8. Shifting Time", "shifting_time", "0", False, False),
            ("9. Turn Round Time (Hrs)-Total=(3+4+5+6)", "trt_total", "#,##0.00", True, False),
            ("    a) Turn Round Time (Hrs)-Port a/c=(3(a)+4+5(a)+6(a)+7+8)", "trt_port", "#,##0.00", False, False),
            ("    b) Turn Round Time (Hrs)-N.P a/c=(3(b)+5(b)+6(b))", "trt_non_port", "#,##0.00", False, False),
        ]

        curr_r = 5
        for label, key, num_fmt, is_bold, is_yellow in rows_def:
            item_cell = ws.cell(row=curr_r, column=1, value=label)
            item_cell.font = bold_font if is_bold else norm_font
            item_cell.alignment = left_align
            item_cell.border = thin_border
            if is_yellow:
                item_cell.fill = yellow_fill

            liq_val = data["liquid_bulk"].get(key, 0)
            liq_cell = ws.cell(row=curr_r, column=2, value=liq_val)
            liq_cell.font = bold_font if is_bold else norm_font
            liq_cell.alignment = right_align
            liq_cell.number_format = num_fmt
            liq_cell.border = thin_border
            if is_yellow:
                liq_cell.fill = yellow_fill

            all_val = data["all_vessels"].get(key, 0)
            all_cell = ws.cell(row=curr_r, column=3, value=all_val)
            all_cell.font = bold_font if is_bold else norm_font
            all_cell.alignment = right_align
            all_cell.number_format = num_fmt
            all_cell.border = thin_border

            curr_r += 1


        ws.column_dimensions["A"].width = 56
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Report-7_Physical_Perf_{fin_year}_{month}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

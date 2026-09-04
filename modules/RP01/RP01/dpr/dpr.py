from .. import bp

from flask import (
    render_template,
    session,
    redirect,
    url_for,
    request,
    jsonify,
    send_file,
)
from functools import wraps
from datetime import datetime, time, timedelta
from io import BytesIO
from database import get_db, get_cursor, get_user_permissions

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

MODULE_CODE = 'RP01'


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def get_perms():
    if session.get('is_admin'):
        return {'can_read': 1, 'can_add': 1, 'can_edit': 1, 'can_delete': 1}
    return get_user_permissions(session.get('user_id'), MODULE_CODE)


def _parse_date(date_str):
    if date_str:
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                pass
    return datetime.now().date()


def _get_window(selected_date):
    """
    7 AM to 7 AM daily window logic:
    If selected_date is 28-08-2026, window is:
    window_start = 27-08-2026 07:00:00
    window_end   = 28-08-2026 07:00:00
    """
    window_end = datetime.combine(selected_date, time(7, 0, 0))
    window_start = window_end - timedelta(days=1)
    return window_start, window_end


def _get_month_window(selected_date, window_end):
    month_start = datetime.combine(selected_date.replace(day=1), time(7, 0, 0))
    return month_start, window_end


def _get_fy_info(selected_date):
    if selected_date.month >= 4:
        fy_start_year = selected_date.year
    else:
        fy_start_year = selected_date.year - 1
    fy_end_short = (fy_start_year + 1) % 100
    fy_label = f"{fy_start_year}-{fy_end_short:02d}"
    fy_start_dt = datetime(fy_start_year, 4, 1, 7, 0, 0)
    return fy_label, fy_start_dt


def _fmt_dt_compact(dt_val):
    """Format datetime as DDMMYYYY HHMM e.g. 24082026 1800"""
    if not dt_val:
        return ''
    if isinstance(dt_val, str):
        s = dt_val.strip()
        for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d-%m-%Y %H:%M'):
            try:
                dt_val = datetime.strptime(s[:16], fmt[:len(s)])
                break
            except Exception:
                pass
    if isinstance(dt_val, datetime):
        return dt_val.strftime('%d-%m-%Y %H:%M')
    return str(dt_val)



def _fmt_eta(dt_val):
    """Format ETA as 20-Aug-26 14.0"""
    if not dt_val:
        return ''
    if isinstance(dt_val, str):
        s = dt_val.strip()
        for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d-%m-%Y %H:%M'):
            try:
                dt_val = datetime.strptime(s[:16], fmt[:len(s)])
                break
            except Exception:
                pass
    if isinstance(dt_val, datetime):
        return dt_val.strftime('%d-%b-%y %H.%M')
    return str(dt_val)


def _parse_entry_dt(entry_date, from_time=None):
    """Safely parse entry_date (text or date) + optional from_time into a datetime object without SQL casting errors."""
    if not entry_date:
        return None

    base_dt = None
    if isinstance(entry_date, datetime):
        base_dt = entry_date
    elif hasattr(entry_date, 'year') and hasattr(entry_date, 'month') and not hasattr(entry_date, 'hour'):
        base_dt = datetime.combine(entry_date, time(0, 0))
    else:
        s = str(entry_date).strip()
        for fmt in (
            '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M',
            '%d-%m-%Y %H:%M:%S', '%d-%m-%Y %H:%M', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M',
            '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%d.%m.%Y'
        ):
            try:
                base_dt = datetime.strptime(s, fmt)
                break
            except Exception:
                continue
        if not base_dt:
            try:
                base_dt = datetime.fromisoformat(s)
            except Exception:
                try:
                    d = datetime.strptime(s[:10], '%Y-%m-%d').date()
                    base_dt = datetime.combine(d, time(0, 0))
                except Exception:
                    return None

    if from_time:
        try:
            parts = str(from_time).strip().split(':')
            if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                return datetime.combine(base_dt.date(), time(int(parts[0]), int(parts[1])))
        except Exception:
            pass

    return base_dt


def _load_vessel_cargo_map(cur):
    """Load cargo master mapping dynamically from database category columns across all 3 tables."""
    master_map = {}

    def resolve_category(cat_val, fallback_name=""):
        s = str(cat_val or '').strip().upper()
        if not s:
            s = str(fallback_name or '').strip().upper()
        if 'EDIBLE' in s:
            return 'Edible Oil'
        if 'POL' in s:
            return 'POL'
        if 'CHEM' in s:
            return 'Chemical'
        return 'Other liquid'

    try:
        # 1. vessel_cargo (master table) -> check cargo_sub_category_2 FIRST, then cargo_category, cargo_type
        cur.execute("""
            SELECT LOWER(TRIM(cargo_name)) AS norm_name, cargo_sub_category_2, cargo_category, cargo_type
            FROM vessel_cargo
            WHERE cargo_name IS NOT NULL AND TRIM(cargo_name) != ''
        """)
        for r in cur.fetchall():
            primary = r.get('cargo_sub_category_2') or r.get('cargo_category') or r.get('cargo_type')
            master_map[r['norm_name']] = resolve_category(primary, r['norm_name'])

        # 2. mis_vessel_master -> use new_cat, category1, category
        cur.execute("""
            SELECT LOWER(TRIM(cargo)) AS norm_name, new_cat, category1, category
            FROM mis_vessel_master
            WHERE cargo IS NOT NULL AND TRIM(cargo) != ''
        """)
        for r in cur.fetchall():
            cn = r['norm_name']
            if cn not in master_map:
                primary = r.get('new_cat') or r.get('category1') or r.get('category')
                master_map[cn] = resolve_category(primary, cn)

        # 3. mis_history -> use cargo_sub_category_2, cargo_category, cargo_type
        cur.execute("""
            SELECT LOWER(TRIM(cargo_name)) AS norm_name, cargo_sub_category_2, cargo_category, cargo_type
            FROM mis_history
            WHERE cargo_name IS NOT NULL AND TRIM(cargo_name) != ''
        """)
        for r in cur.fetchall():
            cn = r['norm_name']
            if cn not in master_map:
                primary = r.get('cargo_sub_category_2') or r.get('cargo_category') or r.get('cargo_type')
                master_map[cn] = resolve_category(primary, cn)

        return master_map
    except Exception:
        return master_map


def _categorize_cargo(cargo_name, cargo_sub2=None, cargo_cat=None, cargo_sub=None, vc_map=None):
    """
    Map cargo to Section B categories: Edible Oil, Other liquid, Chemical, POL.
    In mis_history and master tables, cargo_sub_category_2 has the exact specific bucket
    (CHEMICAL, EDIBLE OIL, FARM LIQUIDS, OTHER LIQUID, POL).
    FARM LIQUIDS and OTHER LIQUID combine into 'Other liquid'.
    """
    # 1. First priority: explicit category fields (cargo_sub2 checked first, e.g. cargo_sub_category_2 or new_cat)
    for val in (cargo_sub2, cargo_cat, cargo_sub):
        if not val:
            continue
        s = str(val).strip().upper()
        if 'EDIBLE' in s:
            return 'Edible Oil'
        if 'CHEM' in s:
            return 'Chemical'
        if 'POL' in s:
            return 'POL'
        if 'FARM' in s or 'OTHER' in s:
            return 'Other liquid'

    # 2. Second priority: master mapping by cargo name
    c_norm = str(cargo_name or '').strip().lower()
    if vc_map and c_norm in vc_map:
        mapped = vc_map[c_norm]
        if mapped in ('Edible Oil', 'Other liquid', 'Chemical', 'POL'):
            return mapped

    # 3. Third priority: cargo name keywords
    s_name = str(cargo_name or '').strip().upper()
    if 'EDIBLE' in s_name or 'PALM' in s_name or 'SOYA' in s_name or 'SUNFLOWER' in s_name or 'CPO' in s_name or 'CDSBO' in s_name:
        return 'Edible Oil'
    if 'POL' in s_name or 'MOTOR SPIRIT' in s_name or 'HSD' in s_name or 'NAPHTHA' in s_name or 'DIESEL' in s_name or 'PETROL' in s_name:
        return 'POL'
    if 'CHEM' in s_name or 'BENZENE' in s_name or 'HEXANE' in s_name or 'METHANOL' in s_name or 'ETHANOL' in s_name or 'ACETONE' in s_name:
        return 'Chemical'

    return 'Other liquid'


def _get_section_a(cur, window_start, window_end):
    """
    Section A: Vessel handling at JJLTPL
    Shows vessels alongside or active at berth slots during 7 AM to 7 AM window.
    A vessel stays on berth if alongside <= window_end AND (cast_off IS NULL OR cast_off > window_end).
    If cast_off <= window_end, the vessel has cast off and is removed from Section A.

    Status logic:
      - If ANY parcel of the vessel is working / in progress: Status = "Operations"
      - If ALL parcels are completed (and no cast_off yet): Status = "Completed"
    """
    cur.execute("""
        SELECT
            vh.id AS vh_id,
            vh.berth_name AS berth,
            vh.vessel_name,
            vh.via_number AS via,
            po.id AS parcel_op_id,
            po.parcel_ids,
            COALESCE(po.cargo_name, vh.cargo_type, 'Liquid Bulk') AS cargo,
            po.quantity AS op_qty,
            po.start_dt,
            po.end_dt,
            po.expected_start,
            COALESCE(po.expected_flow_rate, 0) AS expected_flow_rate,
            lh.alongside_datetime,
            lh.cast_off_datetime
        FROM vcn_header vh
        JOIN ldud_header lh ON lh.vcn_id = vh.id
        LEFT JOIN ldud_parcel_ops po ON po.ldud_id = lh.id
        WHERE (vh.berth_name LIKE '%%LB%%' OR vh.berth_name LIKE '%%03%%' OR vh.berth_name LIKE '%%04%%')
          AND NULLIF(lh.alongside_datetime,'') IS NOT NULL
          AND NULLIF(lh.alongside_datetime,'')::timestamp <= %s
          AND (
              NULLIF(lh.cast_off_datetime,'') IS NULL
              OR NULLIF(lh.cast_off_datetime,'')::timestamp > %s
          )
        ORDER BY vh.berth_name, vh.id, po.id
    """, (window_end, window_end))

    raw_rows = cur.fetchall()

    grouped = {}
    vessel_order = []

    for r in raw_rows:
        vh_id = r['vh_id']
        berth_raw = (r.get('berth') or 'LB03').replace('-', '').strip()
        v_name = (r.get('vessel_name') or '').strip()
        key = (vh_id, berth_raw, v_name)

        if key not in grouped:
            grouped[key] = {
                'vh_id': vh_id,
                'berth': berth_raw,
                'vessel_name': v_name,
                'cargo_list': [],
                'total_bl_qty': 0.0,
                'commenced_times': [],
                'completed_times': [],
                'parcel_statuses': [],
            }
            vessel_order.append(key)

        g = grouped[key]

        c_name = (r.get('cargo') or '').strip()
        if c_name and c_name not in g['cargo_list']:
            g['cargo_list'].append(c_name)

        p_qty = 0.0
        if r.get('op_qty'):
            try:
                p_qty = float(str(r.get('op_qty')).replace(',', '').strip())
            except Exception:
                p_qty = 0.0
        elif r.get('parcel_ids'):
            try:
                cur.execute("""
                    SELECT SUM(q.qty) AS tot FROM (
                        SELECT NULLIF(regexp_replace(COALESCE(quantity, '0'), '[^0-9.]', '', 'g'), '')::numeric AS qty FROM vcn_consigners WHERE vcn_id = %s
                        UNION ALL
                        SELECT NULLIF(regexp_replace(COALESCE(quantity, '0'), '[^0-9.]', '', 'g'), '')::numeric AS qty FROM vcn_export_cargo_declaration WHERE vcn_id = %s
                    ) q
                """, (vh_id, vh_id))
                q_row = cur.fetchone()
                if q_row and q_row.get('tot'):
                    p_qty = float(q_row['tot'])
            except Exception:
                p_qty = 0.0

        g['total_bl_qty'] += p_qty

        start_time = _parse_entry_dt(r.get('start_dt'))
        if not start_time and r.get('alongside_datetime'):
            start_time = _parse_entry_dt(r.get('alongside_datetime'))
        if start_time:
            g['commenced_times'].append(start_time)

        comp_time = _parse_entry_dt(r.get('end_dt'))
        exp_start = _parse_entry_dt(r.get('expected_start'))
        pid = r.get('parcel_op_id')
        real_qty = 0.0
        hours = 0.0

        if pid:
            try:
                cur.execute("""
                    SELECT from_time, to_time, COALESCE(quantity, 0) AS q, is_shortclose, remarks
                    FROM lueu_parcel_log
                    WHERE parcel_op_id = %s AND is_deleted IS NOT TRUE
                    ORDER BY entry_date, from_time NULLS LAST, id
                """, (pid,))
                log_rows = cur.fetchall()
                for lr in log_rows:
                    if lr.get('is_shortclose') or 'short' in str(lr.get('remarks') or '').lower():
                        continue
                    real_qty += float(lr.get('q') or 0)
                    try:
                        fh, fm = (int(x) for x in str(lr['from_time']).split(':')[:2])
                        th, tm = (int(x) for x in str(lr['to_time']).split(':')[:2])
                        m_diff = (th * 60 + tm) - (fh * 60 + fm)
                        if m_diff < 0:
                            m_diff += 1440
                        hours += m_diff / 60.0
                    except Exception:
                        pass
            except Exception:
                pass

        remaining = max(p_qty - real_qty, 0.0) if p_qty > 0 else 0.0

        if comp_time or (p_qty > 0 and remaining <= 1e-6):
            g['parcel_statuses'].append('Completed')
            if comp_time:
                g['completed_times'].append(comp_time)
        else:
            g['parcel_statuses'].append('Operations')
            if start_time and hours > 0 and real_qty > 0 and p_qty > 0:
                avg_rate = real_qty / hours
                if avg_rate > 0:
                    comp_time = start_time + timedelta(hours=(p_qty / avg_rate))
            elif exp_start and r.get('expected_flow_rate') and float(r['expected_flow_rate']) > 0 and p_qty > 0:
                comp_time = exp_start + timedelta(hours=(p_qty / float(r['expected_flow_rate'])))
            if comp_time:
                g['completed_times'].append(comp_time)

    vessels = []
    for key in vessel_order:
        g = grouped[key]

        min_commenced = min(g['commenced_times']) if g['commenced_times'] else None
        max_completed = max(g['completed_times']) if g['completed_times'] else None

        if any(st == 'Operations' for st in g['parcel_statuses']):
            v_status = 'Operations'
        else:
            v_status = 'Completed'

        vessels.append({
            'berth': g['berth'],
            'vessel_name': g['vessel_name'],
            'cargo': " / ".join(g['cargo_list']) if g['cargo_list'] else 'Liquid Bulk',
            'bl_quantity': round(g['total_bl_qty'], 3),
            'discharge_commenced': _fmt_dt_compact(min_commenced),
            'discharge_completed': _fmt_dt_compact(max_completed),
            'status': v_status,
        })


    # Ensure 4 columns as in template (LB03, LB04, LB03, LB04)
    result = []
    default_berths = ['LB03', 'LB04', 'LB03', 'LB04']
    for idx in range(4):
        if idx < len(vessels):
            result.append(vessels[idx])
        else:
            result.append({
                'berth': default_berths[idx],
                'vessel_name': '',
                'cargo': '',
                'bl_quantity': 0.0,
                'discharge_commenced': '',
                'discharge_completed': '',
                'status': ''
            })

    return result




def _get_section_b(cur, selected_date, window_start, window_end):
    """
    Section B: Cargo Handled
    Row categories: Edible Oil, Phosphoric, Lube Oil, Chemicals, POL
    Rows:
    1. Balance on Board to be unloaded/Loaded
    2. Cargo discharged during Day
    3. Cargo discharge in this Month
    4. Cum handled FY 2026-27
    5. Cum handled FY 2025-26
    6. Cum handled FY 2024-25
    7. Cum Loading/unloading till date
    """
    vc_map = _load_vessel_cargo_map(cur)
    categories = ['Edible Oil', 'Other liquid', 'Chemical', 'POL']
    fy_label, _ = _get_fy_info(selected_date)
    month_start, _ = _get_month_window(selected_date, window_end)
    month_name = selected_date.strftime('%b %Y')

    grid = {
        'balance_on_board': {c: 0.0 for c in categories},
        'day_discharged': {c: 0.0 for c in categories},
        'month_discharged': {c: 0.0 for c in categories},
        'cum_fy_curr': {c: 0.0 for c in categories},
        'cum_fy_2025_26': {c: 0.0 for c in categories},
        'cum_fy_2024_25': {c: 0.0 for c in categories},
        'cum_till_date': {c: 0.0 for c in categories},
    }

    # 1. Balance on Board — ONLY vessels currently on berth (LB-03/LB-04, alongside but not cast off)
    cur.execute("""
        SELECT
            po.id AS parcel_op_id,
            COALESCE(po.cargo_name, vh.cargo_type, '') AS cargo_name,
            COALESCE(po.quantity, 0) AS op_qty,
            vc.cargo_category,
            vc.cargo_sub_category,
            vc.cargo_sub_category_2
        FROM ldud_parcel_ops po
        JOIN ldud_header lh ON lh.id = po.ldud_id
        JOIN vcn_header vh ON vh.id = lh.vcn_id
        LEFT JOIN vessel_cargo vc ON LOWER(TRIM(COALESCE(po.cargo_name, ''))) = LOWER(TRIM(vc.cargo_name))
        WHERE COALESCE(lh.is_deleted, FALSE) = FALSE
          AND (vh.berth_name LIKE '%%LB%%' OR vh.berth_name LIKE '%%03%%' OR vh.berth_name LIKE '%%04%%')
          AND NULLIF(lh.alongside_datetime, '') IS NOT NULL
          AND NULLIF(lh.alongside_datetime, '')::timestamp <= %s
          AND (
              NULLIF(lh.cast_off_datetime, '') IS NULL
              OR NULLIF(lh.cast_off_datetime, '')::timestamp > %s
          )
    """, (window_end, window_end))
    active_rows = cur.fetchall()
    for r in active_rows:
        pid = r.get('parcel_op_id')
        c_name = r.get('cargo_name')
        p_qty = 0.0
        if r.get('op_qty'):
            try:
                p_qty = float(str(r.get('op_qty')).replace(',', '').strip())
            except Exception:
                p_qty = 0.0

        real_qty = 0.0
        if pid:
            try:
                cur.execute("""
                    SELECT SUM(COALESCE(quantity, 0)) AS q
                    FROM lueu_parcel_log
                    WHERE parcel_op_id = %s
                      AND COALESCE(is_deleted, FALSE) = FALSE
                      AND COALESCE(is_shortclose, FALSE) = FALSE
                      AND LOWER(COALESCE(remarks, '')) NOT LIKE '%%short%%close%%'
                """, (pid,))
                lr = cur.fetchone()
                if lr and lr.get('q'):
                    real_qty = float(lr['q'])
            except Exception:
                pass

        rem = max(p_qty - real_qty, 0.0)
        cat = _categorize_cargo(c_name, r.get('cargo_category'), r.get('cargo_sub_category'), r.get('cargo_sub_category_2'), vc_map=vc_map)
        if cat not in categories:
            cat = 'Other liquid'
        grid['balance_on_board'][cat] += rem

    # 2 & 3. Cargo discharged during Day and Month from lueu_parcel_log (safe Python datetime parsing)
    cur.execute("""
        SELECT
            po.cargo_name,
            COALESCE(log.quantity, 0) AS qty,
            log.entry_date,
            log.from_time,
            log.to_time,
            log.remarks,
            log.is_shortclose
        FROM lueu_parcel_log log
        JOIN ldud_parcel_ops po ON po.id = log.parcel_op_id
        JOIN ldud_header lh ON lh.id = po.ldud_id
        WHERE COALESCE(log.is_deleted, FALSE) = FALSE
          AND COALESCE(log.is_shortclose, FALSE) = FALSE
          AND COALESCE(lh.is_deleted, FALSE) = FALSE
          AND NULLIF(TRIM(log.entry_date), '') IS NOT NULL
    """)
    for r in cur.fetchall():
        if r.get('is_shortclose') or 'short' in str(r.get('remarks') or '').lower():
            continue
        entry_dt = _parse_entry_dt(r.get('entry_date'), r.get('from_time'))
        if not entry_dt:
            continue
        c_name = r.get('cargo_name')
        q = float(r.get('qty') or 0)
        cat = _categorize_cargo(c_name, vc_map=vc_map)
        if cat not in categories:
            cat = 'Other liquid'

        if window_start <= entry_dt < window_end:
            grid['day_discharged'][cat] += q

        if month_start <= entry_dt < window_end:
            grid['month_discharged'][cat] += q

    # 4. Cum handled FY 2026-27
    # Part 4A: Reconciled MIS data from mis_vessel_master (covers Apr to Jun 2026)
    cur.execute("""
        SELECT
            mvm.cargo,
            mvm.new_cat,
            mvm.category1,
            mvm.category,
            SUM(COALESCE(mvm.quantity, 0)) AS qty
        FROM mis_vessel_master mvm
        WHERE mvm.fin_year = '2026-27'
        GROUP BY mvm.cargo, mvm.new_cat, mvm.category1, mvm.category
    """)
    for r in cur.fetchall():
        q = float(r.get('qty') or 0)
        cat = _categorize_cargo(r.get('cargo'), cargo_sub2=r.get('new_cat'), cargo_cat=r.get('category1'), cargo_sub=r.get('category'), vc_map=vc_map)
        if cat not in categories:
            cat = 'Other liquid'
        grid['cum_fy_curr'][cat] += q

    # Part 4B: Live operational data (from ldud_header, ldud_parcel_ops, lueu_parcel_log)
    # Includes vessels with cast_off_datetime after June (from Jul 1 onwards) up to window_end.
    # Uses vessel_cargo.cargo_sub_category_2 to classify the cargo.
    if selected_date.month < 4:
        live_start = datetime(selected_date.year - 1, 7, 1, 7, 0, 0)
    else:
        live_start = datetime(selected_date.year, 7, 1, 7, 0, 0)

    cur.execute("""
        SELECT
            lh.id AS ldud_id,
            lh.vessel_name,
            lh.cast_off_datetime,
            po.cargo_name,
            vc.cargo_sub_category_2,
            vc.cargo_category,
            vc.cargo_type,
            SUM(COALESCE(log.quantity, 0)) AS qty
        FROM ldud_header lh
        JOIN ldud_parcel_ops po ON po.ldud_id = lh.id
        JOIN lueu_parcel_log log ON log.parcel_op_id = po.id
        LEFT JOIN vessel_cargo vc ON LOWER(TRIM(vc.cargo_name)) = LOWER(TRIM(po.cargo_name))
        WHERE NULLIF(TRIM(lh.cast_off_datetime::text), '') IS NOT NULL
          AND COALESCE(log.is_deleted, FALSE) = FALSE
          AND COALESCE(log.is_shortclose, FALSE) = FALSE
          AND LOWER(COALESCE(log.remarks, '')) NOT LIKE '%%short%%'
          AND COALESCE(lh.is_deleted, FALSE) = FALSE
        GROUP BY
            lh.id,
            lh.vessel_name,
            lh.cast_off_datetime,
            po.cargo_name,
            vc.cargo_sub_category_2,
            vc.cargo_category,
            vc.cargo_type
    """)
    for r in cur.fetchall():
        cast_off_dt = _parse_entry_dt(r.get('cast_off_datetime'))
        if not cast_off_dt:
            continue

        if live_start <= cast_off_dt <= window_end:
            q = float(r.get('qty') or 0)
            if q <= 0:
                continue
            cat = _categorize_cargo(
                r.get('cargo_name'),
                cargo_sub2=r.get('cargo_sub_category_2'),
                cargo_cat=r.get('cargo_category'),
                cargo_sub=r.get('cargo_type'),
                vc_map=vc_map
            )
            if cat not in categories:
                cat = 'Other liquid'
            grid['cum_fy_curr'][cat] += q

    # 5. Cum handled FY 2025-26 (from mis_history)
    cur.execute("""
        SELECT
            mh.cargo_name,
            mh.cargo_sub_category_2,
            mh.cargo_category,
            mh.cargo_type,
            SUM(COALESCE(mh.quantity, 0)) AS qty
        FROM mis_history mh
        WHERE mh.fin_year = '2025-26'
        GROUP BY mh.cargo_name, mh.cargo_sub_category_2, mh.cargo_category, mh.cargo_type
    """)
    for r in cur.fetchall():
        q = float(r.get('qty') or 0)
        cat = _categorize_cargo(r.get('cargo_name'), cargo_sub2=r.get('cargo_sub_category_2'), cargo_cat=r.get('cargo_category'), cargo_sub=r.get('cargo_type'), vc_map=vc_map)
        if cat not in categories:
            cat = 'Other liquid'
        grid['cum_fy_2025_26'][cat] += q

    # 6. Cum handled FY 2024-25 (from mis_history)
    cur.execute("""
        SELECT
            mh.cargo_name,
            mh.cargo_sub_category_2,
            mh.cargo_category,
            mh.cargo_type,
            SUM(COALESCE(mh.quantity, 0)) AS qty
        FROM mis_history mh
        WHERE mh.fin_year = '2024-25'
        GROUP BY mh.cargo_name, mh.cargo_sub_category_2, mh.cargo_category, mh.cargo_type
    """)
    for r in cur.fetchall():
        q = float(r.get('qty') or 0)
        cat = _categorize_cargo(r.get('cargo_name'), cargo_sub2=r.get('cargo_sub_category_2'), cargo_cat=r.get('cargo_category'), cargo_sub=r.get('cargo_type'), vc_map=vc_map)
        if cat not in categories:
            cat = 'Other liquid'
        grid['cum_fy_2024_25'][cat] += q

    # 7. Cum Loading/unloading till date = FY 2026-27 + FY 2025-26 + FY 2024-25
    for c in categories:
        grid['cum_till_date'][c] = (
            grid['cum_fy_curr'][c] +
            grid['cum_fy_2025_26'][c] +
            grid['cum_fy_2024_25'][c]
        )

    # Compute result rows formatted for Section B
    result_rows = []
    row_specs = [
        ('Balance on Board to be unloaded/Loaded', 'balance_on_board'),
        ('Cargo discharged during Day', 'day_discharged'),
        (f'Cargo discharge in this Month ({month_name})', 'month_discharged'),
        (f'Cum handled FY {fy_label}', 'cum_fy_curr'),
        ('Cum handled FY 2025-26', 'cum_fy_2025_26'),
        ('Cum handled FY 2024-25', 'cum_fy_2024_25'),
        ('Cum Loading/unloading till date', 'cum_till_date'),
    ]

    for label, key in row_specs:
        row_dict = {'particulars': label}
        row_total = 0.0
        for cat in categories:
            val = round(grid[key].get(cat, 0.0), 3)
            row_dict[cat] = val
            row_total += val
        row_dict['total'] = round(row_total, 3)
        result_rows.append(row_dict)

    return result_rows


def _get_section_c(cur, selected_date):
    """
    Section C: Vessels Expected
    Dynamic fetch EXCLUSIVELY from EV01 module (expected_vessels table).
    Shows all active / pending expected vessels from EV01.
    Columns: Vessel, Cargo, Berth, B/L Qty, EXIM, Agent, ETA, Remarks
    """
    cur.execute("""
        SELECT
            vessel_name,
            COALESCE(cargo_name, '') AS cargo,
            COALESCE(berth_name, 'LB03/04') AS berth,
            quantity AS bl_qty_str,
            'Import' AS exim,
            COALESCE(agents, '') AS agent,
            eta,
            remarks
        FROM expected_vessels
        WHERE (doc_status IS NULL OR LOWER(doc_status) NOT LIKE '%%closed%%')
        ORDER BY eta ASC NULLS LAST, id
    """)

    ev_rows = cur.fetchall()

    vessels = []
    seen = set()

    for r in ev_rows:
        v_name = (r.get('vessel_name') or '').strip()
        if not v_name or v_name.upper() in seen:
            continue
        seen.add(v_name.upper())

        qty = 0.0
        if r.get('bl_qty_str'):
            try:
                raw_str = str(r.get('bl_qty_str')).replace(',', ' ').strip()
                parts = [float(p) for p in raw_str.split() if p.replace('.', '', 1).isdigit()]
                qty = sum(parts) if parts else 0.0
            except Exception:
                qty = 0.0

        vessels.append({
            'vessel': v_name,
            'cargo': r.get('cargo') or '',
            'berth': r.get('berth') or 'LB03/04',
            'bl_qty': round(qty, 3),
            'exim': r.get('exim') or 'Import',
            'agent': r.get('agent') or '',
            'eta': _fmt_eta(r.get('eta')),
            'remarks': (r.get('remarks') or '').strip(),
        })

    return vessels


def _get_section_d(cur, selected_date):
    """
    Section D: Vessels Handled / operations / declared in [Selected Month Year]
    Dynamic query combining:
    1. Vessels currently running (in operation / alongside) on the selected date window
    2. Vessels completed in the selected month
    Columns: Sr No, Vessel, Cargo, Customer, Quantity, Berth, Agent, Remarks
    """
    window_start, window_end = _get_window(selected_date)
    month_start = datetime.combine(selected_date.replace(day=1), time(7, 0, 0))
    if selected_date.month == 12:
        next_month_start = datetime(selected_date.year + 1, 1, 1, 7, 0, 0)
    else:
        next_month_start = datetime(selected_date.year, selected_date.month + 1, 1, 7, 0, 0)

    month_str = selected_date.strftime('%b-%y')      # e.g. Aug-26
    month_str_long = selected_date.strftime('%b-%Y') # e.g. Aug-2026
    month_name = selected_date.strftime('%B %Y')     # e.g. August 2026

    vessels = []
    seen = set()
    sr = 1

    # 1. RUNNING VESSELS on selected date window
    cur.execute("""
        SELECT
            vh.id AS vh_id,
            vh.vessel_name,
            COALESCE(po.cargo_name, vh.cargo_type, 'Liquid Bulk') AS cargo,
            COALESCE(vh.vessel_agent_name, '') AS agent,
            COALESCE(vh.berth_name, 'LB-03') AS berth,
            lh.alongside_datetime,
            lh.cast_off_datetime
        FROM vcn_header vh
        JOIN ldud_header lh ON lh.vcn_id = vh.id
        LEFT JOIN ldud_parcel_ops po ON po.ldud_id = lh.id
        WHERE NULLIF(lh.alongside_datetime,'') IS NOT NULL
          AND NULLIF(lh.alongside_datetime,'')::timestamp <= %s
          AND (NULLIF(lh.cast_off_datetime,'') IS NULL OR NULLIF(lh.cast_off_datetime,'')::timestamp > %s)
        ORDER BY vh.berth_name, vh.id
    """, (window_end, window_start))

    running_rows = cur.fetchall()
    for r in running_rows:
        v_name = (r.get('vessel_name') or '').strip()
        if not v_name or v_name.upper() in seen:
            continue
        seen.add(v_name.upper())

        vh_id = r.get('vh_id')
        customer_names = []
        qty = 0.0

        if vh_id:
            try:
                cur.execute("""
                    SELECT STRING_AGG(DISTINCT NULLIF(TRIM(c.consigner_name), ''), ', ') AS custs,
                           SUM(NULLIF(regexp_replace(COALESCE(c.quantity, '0'), '[^0-9.]', '', 'g'), '')::numeric) AS tot
                    FROM (
                        SELECT consigner_name, quantity FROM vcn_consigners WHERE vcn_id = %s
                        UNION ALL
                        SELECT consigner_name, quantity FROM vcn_export_cargo_declaration WHERE vcn_id = %s
                    ) c
                """, (vh_id, vh_id))
                c_row = cur.fetchone()
                if c_row:
                    if c_row.get('custs'):
                        customer_names.append(c_row['custs'])
                    if c_row.get('tot'):
                        qty = float(c_row['tot'])
            except Exception:
                pass

        try:
            cur.execute("""
                SELECT SUM(COALESCE(log.quantity, 0)) AS log_qty
                FROM lueu_parcel_log log
                JOIN ldud_parcel_ops po ON po.id = log.parcel_op_id
                JOIN ldud_header lh ON lh.id = po.ldud_id
                WHERE lh.vcn_id = %s AND log.is_deleted IS NOT TRUE
            """, (vh_id,))
            log_r = cur.fetchone()
            if log_r and log_r.get('log_qty') and float(log_r['log_qty']) > 0:
                qty = float(log_r['log_qty'])
        except Exception:
            pass

        has_cast_off = bool(r.get('cast_off_datetime') and str(r.get('cast_off_datetime')).strip())
        has_alongside = bool(r.get('alongside_datetime') and str(r.get('alongside_datetime')).strip())
        display_name = f"{v_name} (In progress)" if (has_alongside and not has_cast_off) else v_name

        vessels.append({
            'sr_no': sr,
            'vessel': display_name,
            'cargo': r.get('cargo') or '',
            'customer': ', '.join(customer_names) if customer_names else 'Multiple Customers',
            'quantity': round(qty, 3),
            'berth': r.get('berth') or 'LB-03',
            'agent': r.get('agent') or '',
            'remarks': '',
            'is_running': not has_cast_off,
        })
        sr += 1

    # 2. LIVE COMPLETED VESSELS in selected month
    cur.execute("""
        SELECT
            vh.id AS vh_id,
            vh.vessel_name,
            COALESCE(po.cargo_name, vh.cargo_type, 'Liquid Bulk') AS cargo,
            COALESCE(vh.vessel_agent_name, '') AS agent,
            COALESCE(vh.berth_name, 'LB-03') AS berth
        FROM vcn_header vh
        JOIN ldud_header lh ON lh.vcn_id = vh.id
        LEFT JOIN ldud_parcel_ops po ON po.ldud_id = lh.id
        WHERE NULLIF(lh.cast_off_datetime,'') IS NOT NULL
          AND NULLIF(lh.cast_off_datetime,'')::timestamp >= %s
          AND NULLIF(lh.cast_off_datetime,'')::timestamp < %s
        ORDER BY lh.cast_off_datetime ASC
    """, (month_start, next_month_start))

    completed_rows = cur.fetchall()
    for r in completed_rows:
        v_name = (r.get('vessel_name') or '').strip()
        if not v_name or v_name.upper() in seen:
            continue
        seen.add(v_name.upper())

        vh_id = r.get('vh_id')
        customer_names = []
        qty = 0.0

        if vh_id:
            try:
                cur.execute("""
                    SELECT STRING_AGG(DISTINCT NULLIF(TRIM(c.consigner_name), ''), ', ') AS custs,
                           SUM(NULLIF(regexp_replace(COALESCE(c.quantity, '0'), '[^0-9.]', '', 'g'), '')::numeric) AS tot
                    FROM (
                        SELECT consigner_name, quantity FROM vcn_consigners WHERE vcn_id = %s
                        UNION ALL
                        SELECT consigner_name, quantity FROM vcn_export_cargo_declaration WHERE vcn_id = %s
                    ) c
                """, (vh_id, vh_id))
                c_row = cur.fetchone()
                if c_row:
                    if c_row.get('custs'):
                        customer_names.append(c_row['custs'])
                    if c_row.get('tot'):
                        qty = float(c_row['tot'])
            except Exception:
                pass

        rem_val = (r.get('remarks') or '').strip()
        if rem_val.lower() == 'completed':
            rem_val = ''

        vessels.append({
            'sr_no': sr,
            'vessel': v_name,
            'cargo': r.get('cargo') or '',
            'customer': ', '.join(customer_names) if customer_names else 'Multiple Customers',
            'quantity': round(qty, 3),
            'berth': r.get('berth') or 'LB-03',
            'agent': r.get('agent') or '',
            'remarks': rem_val,
        })
        sr += 1

    # 3. HISTORICAL COMPLETED VESSELS from mis_vessel_master in selected month
    cur.execute("""
        SELECT
            m.vessel_name,
            m.cargo,
            m.consigner AS customer,
            m.quantity,
            m.berth_no AS berth,
            m.agent,
            m.remarks
        FROM mis_vessel_master m
        WHERE m.month = %s OR m.month = %s
        ORDER BY m.sr_no NULLS LAST, m.id
    """, (month_str, month_str_long))

    hist_rows = cur.fetchall()
    for r in hist_rows:
        v_name = (r.get('vessel_name') or '').strip()
        if not v_name or v_name.upper() in seen:
            continue
        seen.add(v_name.upper())

        h_rem = (r.get('remarks') or '').strip()
        if h_rem.lower() == 'completed':
            h_rem = ''

        vessels.append({
            'sr_no': sr,
            'vessel': v_name,
            'cargo': r.get('cargo') or '',
            'customer': r.get('customer') or '',
            'quantity': float(r.get('quantity') or 0.0),
            'berth': r.get('berth') or 'LB-03',
            'agent': r.get('agent') or '',
            'remarks': h_rem,
        })
        sr += 1

    total_qty = sum(v['quantity'] for v in vessels)

    return {
        'month_name': month_name,
        'vessels': vessels,
        'total_quantity': round(total_qty, 3)
    }



def _get_dpr_payload(selected_date):
    window_start, window_end = _get_window(selected_date)
    reporting_date = selected_date - timedelta(days=1)

    conn = get_db()
    try:
        cur = get_cursor(conn)
        section_a = _get_section_a(cur, window_start, window_end)
        section_b = _get_section_b(cur, selected_date, window_start, window_end)
        section_c = _get_section_c(cur, selected_date)
        section_d = _get_section_d(cur, selected_date)
    finally:
        conn.close()

    return {
        'report_date': selected_date.strftime('%Y-%m-%d'),
        'report_date_display': selected_date.strftime('%d %B %Y'),
        'reporting_date_display': reporting_date.strftime('%d %B %Y'),
        'window_start_display': window_start.strftime('%d-%m-%Y %H:%M'),
        'window_end_display': window_end.strftime('%d-%m-%Y %H:%M'),
        'section_a': section_a,
        'section_b': section_b,
        'section_c': section_c,
        'section_d': section_d,
    }


# ══════════════════════════════════════════════════════════════════
# Routes
# ══════════════════════════════════════════════════════════════════

@bp.route('/module/RP01/dpr/')
@login_required
def dpr_page():
    perms = get_perms()
    if not perms.get('can_read'):
        return render_template('no_access.html'), 403
    return render_template('dpr/dpr.html', permissions=perms)


@bp.route('/api/module/RP01/dpr/data')
@login_required
def dpr_data():
    selected_date = _parse_date(request.args.get('date'))
    try:
        return jsonify(_get_dpr_payload(selected_date))
    except Exception as e:
        import traceback, sys
        tb = traceback.format_exc()
        print(f"[DPR ERROR] {e}\n{tb}", file=sys.stderr, flush=True)
        return jsonify({'error': str(e), 'traceback': tb}), 500


def _populate_dpr_sheet(ws, payload):
    FONT_NAME = "Arial"
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_header = PatternFill("solid", fgColor="BCD6EE")
    fill_section = PatternFill("solid", fgColor="DDEBF7")
    fill_month = PatternFill("solid", fgColor="FFF2A8")
    fill_total = PatternFill("solid", fgColor="FCE0CD")
    fill_white = PatternFill("solid", fgColor="FFFFFF")
    fill_title = PatternFill("solid", fgColor="1F4E78")

    font_title = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
    font_header = Font(name=FONT_NAME, bold=True, size=10)
    font_section = Font(name=FONT_NAME, bold=True, size=10)
    font_normal = Font(name=FONT_NAME, size=10)
    font_value = Font(name=FONT_NAME, size=10, color="1F4E78")
    font_total = Font(name=FONT_NAME, bold=True, size=10)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    NUM_FMT = '#,##0.000;(#,##0.000);"-"'

    def set_cell(coord, value, font=font_normal, fill=None, align=center, fmt=None):
        c = ws[coord]
        c.value = value
        c.font = font
        c.border = border
        c.alignment = align
        if fill:
            c.fill = fill
        if fmt:
            c.number_format = fmt
        return c

    def merge(rng, value=None, font=font_normal, fill=None, align=center, fmt=None):
        ws.merge_cells(rng)
        top_left = rng.split(":")[0]
        set_cell(top_left, value, font=font, fill=fill, align=align, fmt=fmt)
        for row in ws[rng]:
            for cell in row:
                cell.border = border
                if fill:
                    cell.fill = fill

    # Column widths (A to H)
    widths = {
        "A": 36, "B": 22, "C": 18, "D": 18,
        "E": 18, "F": 18, "G": 22, "H": 26
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title Bar (A1:H1)
    merge("A1:H1", "DAILY REPORT - JSW JNPT LIQUID TERMINAL PRIVATE LIMITED (JJLTPL)",
          font=font_title, fill=fill_title, align=center)

    # Meta dates strip (A3:H3)
    set_cell("A3", "Report Date", font=font_header, fill=fill_header, align=left)
    set_cell("B3", payload['report_date_display'], font=font_value, fill=fill_white, align=center)
    merge("C3:E3", "", fill=fill_white)
    set_cell("F3", "Reporting Date", font=font_header, fill=fill_header, align=left)
    merge("G3:H3", payload['reporting_date_display'], font=font_value, fill=fill_white, align=center)

    r = 5
    # Section A
    merge(f"A{r}:H{r}", "Section A: Vessel handling at JJLTPL", font=font_section, fill=fill_section, align=left)

    r += 1
    sec_a_fields = [
        ("Berth", [v['berth'] for v in payload['section_a']]),
        ("Vessel Name", [v['vessel_name'] for v in payload['section_a']]),
        ("Cargo", [v['cargo'] for v in payload['section_a']]),
        ("BL Quantity", [v['bl_quantity'] for v in payload['section_a']]),
        ("Discharge Commenced", [v['discharge_commenced'] for v in payload['section_a']]),
        ("Discharge Completed (ETC)", [v['discharge_completed'] for v in payload['section_a']]),
        ("Status", [v['status'] for v in payload['section_a']]),
    ]

    for label, vals in sec_a_fields:
        set_cell(f"A{r}", label, font=font_header, fill=fill_header, align=left)
        for idx in range(7):  # Columns B through H
            col_let = get_column_letter(2 + idx)
            val = vals[idx] if idx < len(vals) else None
            is_bl = (label == "BL Quantity") and isinstance(val, (int, float))
            set_cell(f"{col_let}{r}", val, font=font_value if val is not None else font_normal,
                     fill=fill_white, align=right if is_bl else center,
                     fmt=NUM_FMT if is_bl else None)
        r += 1

    # Section B
    r += 1
    merge(f"A{r}:H{r}", "Section B: Cargo Handled", font=font_section, fill=fill_section, align=left)

    r += 1
    b_headers = ["Particulars", "Edible Oil", "Other liquid", "Chemical", "POL", "Total"]
    for i, h in enumerate(b_headers):
        col_let = get_column_letter(1 + i)
        set_cell(f"{col_let}{r}", h, font=font_header, fill=fill_header, align=left if i == 0 else right)
    set_cell(f"G{r}", None, fill=fill_white)
    set_cell(f"H{r}", None, fill=fill_white)

    for row_data in payload['section_b']:
        r += 1
        is_total = "till date" in row_data['particulars'].lower()
        is_month = "month" in row_data['particulars'].lower()
        row_fill = fill_total if is_total else (fill_month if is_month else fill_white)
        row_font = font_total if is_total else font_normal

        set_cell(f"A{r}", row_data['particulars'], font=row_font, fill=row_fill, align=left)

        cats = ["Edible Oil", "Other liquid", "Chemical", "POL", "total"]
        for idx, cat in enumerate(cats):
            col_let = get_column_letter(2 + idx)
            val = row_data.get(cat, 0.0)
            cell_font = font_total if (cat == "total" or is_total) else font_value
            cell_fill = fill_total if (cat == "total" or is_total) else row_fill
            set_cell(f"{col_let}{r}", val, font=cell_font, fill=cell_fill, align=right, fmt=NUM_FMT)

        set_cell(f"G{r}", None, fill=fill_white)
        set_cell(f"H{r}", None, fill=fill_white)

    # Section C
    r += 2
    merge(f"A{r}:H{r}", "Section C: Vessels Expected", font=font_section, fill=fill_section, align=left)

    r += 1
    c_headers = ["Vessel", "Cargo", "Berth", "B/L Qty", "EXIM", "Agent", "ETA", "Remarks"]
    for i, h in enumerate(c_headers):
        col_let = get_column_letter(1 + i)
        set_cell(f"{col_let}{r}", h, font=font_header, fill=fill_header, align=right if i == 3 else (left if i in (0, 1, 5, 7) else center))

    if payload['section_c']:
        for v in payload['section_c']:
            r += 1
            vals = [v['vessel'], v['cargo'], v['berth'], v['bl_qty'], v['exim'], v['agent'], v['eta'], v['remarks']]
            for i, val in enumerate(vals):
                col_let = get_column_letter(1 + i)
                is_num = (i == 3 and isinstance(val, (int, float)))
                set_cell(f"{col_let}{r}", val, font=font_value, fill=fill_white,
                         align=right if is_num else (left if i in (0, 1, 5, 7) else center),
                         fmt=NUM_FMT if is_num else None)
    else:
        r += 1
        merge(f"A{r}:H{r}", "No expected vessels for selected month", font=font_normal, fill=fill_white, align=center)

    # Section D
    r += 2
    sec_d = payload['section_d']
    merge(f"A{r}:H{r}", f"Section D: Vessels Handled / operations / declared in {sec_d['month_name']}",
          font=font_section, fill=fill_section, align=left)

    r += 1
    d_headers = ["Vessel", "Cargo", "Customer", "Quantity", "Berth", "Agent", "Remarks"]
    for i, h in enumerate(d_headers):
        col_let = get_column_letter(1 + i)
        set_cell(f"{col_let}{r}", h, font=font_header, fill=fill_header, align=right if i == 3 else (left if i in (0, 1, 2, 5, 6) else center))
    set_cell(f"H{r}", None, fill=fill_white)

    if sec_d['vessels']:
        for v in sec_d['vessels']:
            r += 1
            vals = [v['vessel'], v['cargo'], v['customer'], v['quantity'], v['berth'], v['agent'], v['remarks']]
            for i, val in enumerate(vals):
                col_let = get_column_letter(1 + i)
                is_num = (i == 3 and isinstance(val, (int, float)))
                set_cell(f"{col_let}{r}", val, font=font_value, fill=fill_white,
                         align=right if is_num else (left if i in (0, 1, 2, 5, 6) else center),
                         fmt=NUM_FMT if is_num else None)
            set_cell(f"H{r}", None, fill=fill_white)
    else:
        r += 1
        merge(f"A{r}:G{r}", "No vessels handled in selected month", font=font_normal, fill=fill_white, align=center)
        set_cell(f"H{r}", None, fill=fill_white)

    # Total Row for Section D
    r += 1
    merge(f"A{r}:C{r}", "Total", font=font_total, fill=fill_total, align=left)
    set_cell(f"D{r}", sec_d['total_quantity'], font=font_total, fill=fill_total, align=right, fmt=NUM_FMT)
    merge(f"E{r}:H{r}", "", fill=fill_total)

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = True


@bp.route('/api/module/RP01/dpr/export')
@login_required
def dpr_export():
    selected_date = _parse_date(request.args.get('date'))
    dpr_payload = _get_dpr_payload(selected_date)
    bvsa_payload = _get_bvsa_payload(selected_date)

    wb = Workbook()

    # Sheet 1: DPR
    ws_dpr = wb.active
    ws_dpr.title = "DPR"
    _populate_dpr_sheet(ws_dpr, dpr_payload)

    # Sheet 2: BVsA FY 2027
    ws_bvsa = wb.create_sheet(title="BVsA FY 2027")
    _populate_bvsa_sheet(ws_bvsa, bvsa_payload)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Daily_Performance_Report_{dpr_payload['report_date']}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# TAB 2: BVsA FY 2027 (Budget vs Actual)
# Completely independent calculation for the second tab of DPR.
# Uses financial_year_targets for Budget, and mis_vessel_master + live
# ldud_header / parcel ops for Actuals.
# ---------------------------------------------------------------------------

def _classify_bvsa_cargo(cargo_name, cargo_sub2=None, cargo_cat=None, cargo_sub=None):
    """
    Classifies cargo into the 4 DPR Section B categories:
      - 'edible': Edible Oil
      - 'other': Other liquid (combining Lube, Base Oil, Farm Liquids, Phosphoric Acid)
      - 'chemical': Chemical
      - 'pol': POL
    """
    s_sub2 = str(cargo_sub2 or '').strip().upper()
    s_cat = str(cargo_cat or '').strip().upper()
    s_sub = str(cargo_sub or '').strip().upper()
    s_name = str(cargo_name or '').strip().upper()

    for s in (s_sub2, s_cat, s_sub):
        if 'EDIBLE' in s:
            return 'edible'
        if 'CHEM' in s:
            return 'chemical'
        if 'POL' in s:
            return 'pol'
        if 'FARM' in s or 'OTHER' in s or 'PH' in s or 'LUBE' in s:
            return 'other'

    if any(k in s_name for k in ('EDIBLE', 'PALM', 'SOYA', 'SUNFLOWER', 'CPO', 'CDSBO', 'CSFO')):
        return 'edible'
    if any(k in s_name for k in ('POL', 'CBFS', 'FO', 'HSD', 'MOTOR SPIRIT', 'DIESEL', 'NAPHTHA')):
        return 'pol'
    if any(k in s_name for k in ('CHEM', 'ACETONE', 'PHENOL', 'BENZENE', 'ETHANOL', 'METHANOL', 'ACETIC ACID', 'VAM', 'IPA', 'SM', 'MEK', 'NITRIC ACID', 'HEXANE')):
        return 'chemical'

    return 'other'


def _get_bvsa_payload(selected_date):
    """
    Budget vs Actual (BVsA) calculation for FY 2026-27 (FY 2027).
    - Categories: Edible Oil, Other liquid, Chemical, POL
    - Budget source: financial_year_targets (targets JSONB)
    - Actual source: mis_vessel_master (Apr-Jun) + Live ldud_header / parcel ops (Jul onwards up to selected_date)
    """
    conn = get_db()
    try:
        cur = get_cursor(conn)
        window_end = datetime(selected_date.year, selected_date.month, selected_date.day, 7, 0, 0)

        MONTH_SPECS = [
            ('Apr', 'Apr-26', 4, 2026),
            ('May', 'May-26', 5, 2026),
            ('Jun', 'Jun-26', 6, 2026),
            ('Jul', 'Jul-26', 7, 2026),
            ('Aug', 'Aug-26', 8, 2026),
            ('Sep', 'Sep-26', 9, 2026),
            ('Oct', 'Oct-26', 10, 2026),
            ('Nov', 'Nov-26', 11, 2026),
            ('Dec', 'Dec-26', 12, 2026),
            ('Jan', 'Jan-27', 1, 2027),
            ('Feb', 'Feb-27', 2, 2027),
            ('Mar', 'Mar-27', 3, 2027),
        ]

        if selected_date.year == 2026 and selected_date.month >= 4:
            active_idx = selected_date.month - 4
        elif selected_date.year == 2027 and selected_date.month <= 3:
            active_idx = selected_date.month + 8
        elif selected_date.year > 2027 or (selected_date.year == 2027 and selected_date.month > 3):
            active_idx = 11
        else:
            active_idx = 0

        months = []
        for idx, (m_code, m_lbl, m_num, m_yr) in enumerate(MONTH_SPECS):
            months.append({
                'index': idx,
                'code': m_code,
                'label': m_lbl,
                'month_num': m_num,
                'year': m_yr,
                'is_past_or_current': idx <= active_idx,
                'budget': {'edible': 0.0, 'other': 0.0, 'chemical': 0.0, 'pol': 0.0, 'total': 0.0},
                'actual': {'edible': 0.0, 'other': 0.0, 'chemical': 0.0, 'pol': 0.0, 'total': 0.0}
            })

        # 1. Budget from financial_year_targets
        cur.execute("SELECT targets FROM financial_year_targets WHERE financial_year = '2026-27'")
        fy_row = cur.fetchone()
        if fy_row and fy_row.get('targets'):
            raw_t = fy_row['targets']
            if isinstance(raw_t, str):
                import json
                raw_t = json.loads(raw_t)

            target_items = raw_t.get('targets', []) if isinstance(raw_t, dict) else []
            if not target_items and isinstance(raw_t, dict):
                for grp in raw_t.get('cargo_groups', []):
                    target_items.append({
                        'name': grp.get('cargo_sub_category_2') or grp.get('name') or '',
                        'monthly_data': [{'month': mb.get('month'), 'target': mb.get('budget_quantity')} for mb in grp.get('monthly_budget', [])]
                    })

            for item in target_items:
                name = (item.get('name') or '').strip().upper()
                if 'EDIBLE' in name or 'PALM' in name or 'SOYA' in name:
                    b_cat = 'edible'
                elif 'POL' in name:
                    b_cat = 'pol'
                elif 'CHEM' in name:
                    b_cat = 'chemical'
                else:
                    b_cat = 'other'

                for md in item.get('monthly_data', []):
                    m_str = (md.get('month') or '').strip()
                    val = float(md.get('target') or md.get('budget_quantity') or 0.0)
                    for m in months:
                        if m['code'].lower() == m_str[:3].lower():
                            m['budget'][b_cat] += val
                            m['budget']['total'] += val
                            break

        # 2. Actuals from mis_vessel_master (Apr to Jun 2026)
        vessels = []
        sr = 1

        cur.execute("""
            SELECT 
                id, month, vessel_name, cargo, consigner, new_cat, category1, category, quantity
            FROM mis_vessel_master
            WHERE fin_year = '2026-27'
            ORDER BY 
                CASE 
                    WHEN month ILIKE 'Apr%' THEN 1
                    WHEN month ILIKE 'May%' THEN 2
                    WHEN month ILIKE 'Jun%' THEN 3
                    WHEN month ILIKE 'Jul%' THEN 4
                    WHEN month ILIKE 'Aug%' THEN 5
                    WHEN month ILIKE 'Sep%' THEN 6
                    WHEN month ILIKE 'Oct%' THEN 7
                    WHEN month ILIKE 'Nov%' THEN 8
                    WHEN month ILIKE 'Dec%' THEN 9
                    WHEN month ILIKE 'Jan%' THEN 10
                    WHEN month ILIKE 'Feb%' THEN 11
                    WHEN month ILIKE 'Mar%' THEN 12
                    ELSE 99
                END,
                id
        """)
        for r in cur.fetchall():
            m_str = (r.get('month') or '').strip()
            m_obj = next((m for m in months if m['label'].lower() == m_str.lower() or m['code'].lower() == m_str[:3].lower()), None)
            if not m_obj:
                continue

            q = float(r.get('quantity') or 0.0)
            c_name = r.get('cargo')
            cat = _classify_bvsa_cargo(c_name, r.get('new_cat'), r.get('category1'), r.get('category'))

            m_obj['actual'][cat] += q
            m_obj['actual']['total'] += q

            vessels.append({
                'sr_no': sr,
                'month': m_obj['label'],
                'vessel_name': r.get('vessel_name') or '-',
                'cargo': c_name or '-',
                'customer': r.get('consigner') or '-',
                'quantity': q,
                'edible': q if cat == 'edible' else 0.0,
                'other': q if cat == 'other' else 0.0,
                'chemical': q if cat == 'chemical' else 0.0,
                'pol': q if cat == 'pol' else 0.0,
            })
            sr += 1

        # 3. Actuals from Live Operational Data (Jul 1 07:00 onwards up to window_end)
        cur.execute("""
            WITH parcel_consignee AS (
                SELECT 
                    po.id AS po_id,
                    COALESCE(
                        NULLIF(
                            TRIM((
                                SELECT STRING_AGG(DISTINCT TRIM(customer_name), ', ')
                                FROM vcn_cargo_declaration
                                WHERE id::text = po.parcel_ids
                                  AND NULLIF(TRIM(customer_name), '') IS NOT NULL
                            )),
                            ''
                        ),
                        NULLIF(
                            TRIM((
                                SELECT STRING_AGG(DISTINCT TRIM(customer_name), ', ')
                                FROM vcn_cargo_declaration
                                WHERE vcn_id = lh.vcn_id
                                  AND cargo_name = po.cargo_name
                                  AND NULLIF(TRIM(customer_name), '') IS NOT NULL
                            )),
                            ''
                        ),
                        '-'
                    ) AS consignee_name
                FROM ldud_parcel_ops po
                JOIN ldud_header lh 
                    ON lh.id = po.ldud_id
                LEFT JOIN vcn_header vh 
                    ON vh.id = lh.vcn_id
            )
            SELECT
                lh.id AS ldud_id,
                lh.vessel_name,
                lh.cast_off_datetime,
                po.cargo_name,
                pc.consignee_name AS customer,
                vc.cargo_sub_category_2,
                vc.cargo_category,
                vc.cargo_type,
                SUM(COALESCE(log.quantity, 0)) AS qty
            FROM ldud_header lh
            JOIN ldud_parcel_ops po ON po.ldud_id = lh.id
            JOIN parcel_consignee pc ON pc.po_id = po.id
            JOIN lueu_parcel_log log ON log.parcel_op_id = po.id
            LEFT JOIN vessel_cargo vc ON LOWER(TRIM(vc.cargo_name)) = LOWER(TRIM(po.cargo_name))
            WHERE NULLIF(TRIM(lh.cast_off_datetime::text), '') IS NOT NULL
              AND REPLACE(TRIM(lh.cast_off_datetime), 'T', ' ')::timestamp >= '2026-07-01 07:00:00'
              AND REPLACE(TRIM(lh.cast_off_datetime), 'T', ' ')::timestamp <= %s
              AND COALESCE(log.is_deleted, FALSE) = FALSE
              AND COALESCE(log.is_shortclose, FALSE) = FALSE
              AND LOWER(COALESCE(log.remarks, '')) NOT LIKE '%%short%%'
              AND COALESCE(lh.is_deleted, FALSE) = FALSE
            GROUP BY lh.id, lh.vessel_name, lh.cast_off_datetime, po.cargo_name, pc.consignee_name, vc.cargo_sub_category_2, vc.cargo_category, vc.cargo_type
            ORDER BY lh.cast_off_datetime
        """, (window_end,))
        for r in cur.fetchall():
            co_dt = _parse_entry_dt(r.get('cast_off_datetime'))
            if not co_dt:
                continue

            m_obj = next((m for m in months if m['month_num'] == co_dt.month and m['year'] == co_dt.year), None)
            if not m_obj:
                continue

            q = float(r.get('qty') or 0.0)
            c_name = r.get('cargo_name')
            cat = _classify_bvsa_cargo(c_name, r.get('cargo_sub_category_2'), r.get('cargo_category'), r.get('cargo_type'))

            m_obj['actual'][cat] += q
            m_obj['actual']['total'] += q

            vessels.append({
                'sr_no': sr,
                'month': m_obj['label'],
                'vessel_name': r.get('vessel_name') or '-',
                'cargo': c_name or '-',
                'customer': r.get('customer') or 'Live Operation',
                'quantity': q,
                'edible': q if cat == 'edible' else 0.0,
                'other': q if cat == 'other' else 0.0,
                'chemical': q if cat == 'chemical' else 0.0,
                'pol': q if cat == 'pol' else 0.0,
            })
            sr += 1

        # 4. Cumulative calculations and Variances (Variance = Actual - Budget)
        full_year_budget = sum(m['budget']['total'] for m in months)
        cum_b = 0.0
        cum_a = 0.0

        for m in months:
            cum_b += m['budget']['total']
            m['cum_budget'] = cum_b

            if m['is_past_or_current']:
                cum_a += m['actual']['total']
                m['cum_actual'] = cum_a
                m['pct_achieved'] = (m['actual']['total'] / m['budget']['total'] * 100) if m['budget']['total'] > 0 else (100.0 if m['actual']['total'] > 0 else 0.0)
                m['pct_cum_achieved'] = (cum_a / full_year_budget * 100) if full_year_budget > 0 else 0.0
                m['variance'] = cum_a - cum_b
                m['cum_variance'] = cum_a - cum_b
            else:
                m['cum_actual'] = None
                m['pct_achieved'] = None
                m['pct_cum_achieved'] = None
                m['variance'] = None
                m['cum_variance'] = None

        active_m = months[active_idx]
        ytd_budget = active_m['cum_budget']
        ytd_actual = active_m['cum_actual'] or 0.0
        ytd_variance = ytd_actual - ytd_budget
        balance = max(full_year_budget - ytd_actual, 0.0)
        remaining_months = max(0, 11 - active_idx)
        asking_rate = (balance / remaining_months) if remaining_months > 0 else 0.0

        # Category totals for FY & YTD
        fy_summary = {
            'edible': {
                'budget': sum(m['budget']['edible'] for m in months),
                'actual': sum(m['actual']['edible'] for m in months if m['is_past_or_current']),
            },
            'other': {
                'budget': sum(m['budget']['other'] for m in months),
                'actual': sum(m['actual']['other'] for m in months if m['is_past_or_current']),
            },
            'chemical': {
                'budget': sum(m['budget']['chemical'] for m in months),
                'actual': sum(m['actual']['chemical'] for m in months if m['is_past_or_current']),
            },
            'pol': {
                'budget': sum(m['budget']['pol'] for m in months),
                'actual': sum(m['actual']['pol'] for m in months if m['is_past_or_current']),
            },
            'total': {
                'budget': full_year_budget,
                'actual': ytd_actual,
                'balance': balance,
            }
        }
        for k in ('edible', 'other', 'chemical', 'pol'):
            fy_summary[k]['balance'] = max(fy_summary[k]['budget'] - fy_summary[k]['actual'], 0.0)

        # Quarterly breakdown for FY 2026-27
        q1_act = sum(months[i]['actual']['total'] for i in (0, 1, 2) if months[i]['is_past_or_current'])
        q2_act = sum(months[i]['actual']['total'] for i in (3, 4, 5) if months[i]['is_past_or_current'])
        q3_act = sum(months[i]['actual']['total'] for i in (6, 7, 8) if months[i]['is_past_or_current'])
        q4_act = sum(months[i]['actual']['total'] for i in (9, 10, 11) if months[i]['is_past_or_current'])

        # Historical actuals queried dynamically from database
        cur.execute("SELECT fin_year, SUM(quantity) AS tot FROM mis_history GROUP BY fin_year")
        hist_rows = {r['fin_year']: float(r['tot'] or 0.0) for r in cur.fetchall()}
        fy25_tot = hist_rows.get('2024-25', 200912.538)
        fy26_tot = hist_rows.get('2025-26', 1300492.320)

        cur.execute("""
            SELECT month, SUM(quantity) as tot 
            FROM mis_vessel_master 
            WHERE fin_year = '2025-26' 
            GROUP BY month
        """)
        m_26 = {r['month']: float(r['tot'] or 0.0) for r in cur.fetchall()}
        fy26_q1 = sum(m_26.get(k, 0.0) for k in ('Apr-25', 'May-25', 'Jun-25')) or 293549.0
        fy26_q2 = sum(m_26.get(k, 0.0) for k in ('Jul-25', 'Aug-25', 'Sep-25')) or 444804.0
        fy26_q3 = sum(m_26.get(k, 0.0) for k in ('Oct-25', 'Nov-25', 'Dec-25')) or 298730.0
        fy26_q4 = sum(m_26.get(k, 0.0) for k in ('Jan-26', 'Feb-26', 'Mar-26')) or 263409.0

        # Ensure all numeric quantities in BVsA payload have exact 3 decimal places
        for m in months:
            for k in ('edible', 'other', 'chemical', 'pol', 'total'):
                m['budget'][k] = round(float(m['budget'][k]), 3)
                m['actual'][k] = round(float(m['actual'][k]), 3)
            m['cum_budget'] = round(float(m['cum_budget']), 3)
            if m['cum_actual'] is not None:
                m['cum_actual'] = round(float(m['cum_actual']), 3)
            if m['variance'] is not None:
                m['variance'] = round(float(m['variance']), 3)
            if m['cum_variance'] is not None:
                m['cum_variance'] = round(float(m['cum_variance']), 3)

        for v in vessels:
            v['quantity'] = round(float(v['quantity']), 3)
            v['edible'] = round(float(v['edible']), 3)
            v['other'] = round(float(v['other']), 3)
            v['chemical'] = round(float(v['chemical']), 3)
            v['pol'] = round(float(v['pol']), 3)

        full_year_budget = round(float(full_year_budget), 3)
        ytd_budget = round(float(ytd_budget), 3)
        ytd_actual = round(float(ytd_actual), 3)
        ytd_variance = round(float(ytd_variance), 3)
        balance = round(float(balance), 3)
        asking_rate = round(float(asking_rate), 3)

        for k in ('edible', 'other', 'chemical', 'pol'):
            fy_summary[k]['budget'] = round(float(fy_summary[k]['budget']), 3)
            fy_summary[k]['actual'] = round(float(fy_summary[k]['actual']), 3)
            fy_summary[k]['balance'] = round(float(fy_summary[k]['balance']), 3)
        fy_summary['total']['budget'] = full_year_budget
        fy_summary['total']['actual'] = ytd_actual
        fy_summary['total']['balance'] = balance

        q1_act = round(float(q1_act), 3)
        q2_act = round(float(q2_act), 3)
        q3_act = round(float(q3_act), 3)
        q4_act = round(float(q4_act), 3)
        fy25_tot = round(float(fy25_tot), 3)
        fy26_tot = round(float(fy26_tot), 3)

        return {
            'financial_year': '2026-27',
            'report_date': selected_date.isoformat(),
            'active_month_idx': active_idx,
            'active_month_label': active_m['label'],
            'months': months,
            'vessels': vessels,
            'summary': {
                'full_year_budget': full_year_budget,
                'ytd_budget': ytd_budget,
                'ytd_actual': ytd_actual,
                'ytd_variance': ytd_variance,
                'balance': balance,
                'remaining_months': remaining_months,
                'asking_rate': asking_rate,
                'pct_achieved_ytd': (ytd_actual / ytd_budget * 100) if ytd_budget > 0 else 0.0,
                'pct_cum_fy': (ytd_actual / full_year_budget * 100) if full_year_budget > 0 else 0.0,
                'category_summary': fy_summary,
                'quarterly': {
                    'q1': q1_act,
                    'q2': q2_act,
                    'q3': q3_act,
                    'q4': q4_act,
                },
                'historical': {
                    'fy25': fy25_tot,
                    'fy26': fy26_tot,
                    'fy26_q1': fy26_q1,
                    'fy26_q2': fy26_q2,
                    'fy26_q3': fy26_q3,
                    'fy26_q4': fy26_q4,
                }
            }
        }
    finally:
        conn.close()


@bp.route('/api/module/RP01/dpr/bvsa')
@login_required
def dpr_bvsa_data():
    """Returns dynamic BVsA FY 2027 payload."""
    try:
        date_str = request.args.get('date')
        selected_date = _parse_date(date_str)
        return jsonify(_get_bvsa_payload(selected_date))
    except Exception as e:
        import traceback, sys
        tb = traceback.format_exc()
        print(f"[DPR BVsA ERROR] {e}\n{tb}", file=sys.stderr, flush=True)
        return jsonify({'error': str(e), 'traceback': tb}), 500


def _populate_bvsa_sheet(ws, payload):
    """Populates BVsA FY 2027 worksheet matching UI structure, table borders, gap spacing, and colors."""
    FONT_NAME = "Calibri"
    thin_gray = Side(style="thin", color="D9D9D9")
    double_bottom = Side(style="double", color="000000")

    border_cell = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    border_total = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=double_bottom)

    # UI-identical Color Fills
    fill_navy = PatternFill("solid", fgColor="1F4E78")      # .th-blue
    fill_sub_blue = PatternFill("solid", fgColor="2F5597")  # .th-sub-blue
    fill_curr_month = PatternFill("solid", fgColor="D66011")# Active month accent
    fill_soft_green = PatternFill("solid", fgColor="C6EFCE")# .tr-green
    fill_cum_budget = PatternFill("solid", fgColor="E2EFDA")# .tr-cum-budget
    fill_highlight = PatternFill("solid", fgColor="DDEBF7") # .tr-highlight
    fill_white = PatternFill("solid", fgColor="FFFFFF")
    fill_month_alt = PatternFill("solid", fgColor="EAF2F8") # .vessel-month-alt
    fill_amber = PatternFill("solid", fgColor="FFC000")     # Month Total cell

    # Typography
    font_title = Font(name=FONT_NAME, size=10.5, bold=True, color="FFFFFF")
    font_header = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    font_bold = Font(name=FONT_NAME, size=10, bold=True, color="000000")
    font_normal = Font(name=FONT_NAME, size=10, color="000000")
    font_italic = Font(name=FONT_NAME, size=10, italic=True, color="475569")
    font_green_bold = Font(name=FONT_NAME, size=10, bold=True, color="006100")
    font_var_neg = Font(name=FONT_NAME, size=10, bold=True, color="C00000")
    font_var_pos = Font(name=FONT_NAME, size=10, bold=True, color="006100")
    font_floating_sum = Font(name=FONT_NAME, size=11, bold=True, color="1F4E78")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    NUM_FMT = "#,##0.000"

    def set_c(addr, val, font=font_normal, fill=fill_white, align=right, fmt=None, bdr=border_cell):
        c = ws[addr]
        c.value = val
        c.font = font
        c.fill = fill
        c.alignment = align
        c.border = bdr
        if fmt:
            c.number_format = fmt
        return c

    # 1. TOP CARGO PROJECTIONS TABLE
    set_c("B1", "CARGO PROJECTIONS", font=font_title, fill=fill_navy, align=left)
    set_c("C1", None, fill=fill_navy)
    ws.merge_cells("B1:C1")

    for i, m in enumerate(payload['months']):
        col_let = get_column_letter(4 + i) # D to O
        is_curr = (i == payload.get('active_month_idx'))
        f_fill = fill_curr_month if is_curr else fill_navy
        set_c(f"{col_let}1", m['label'], font=font_header, fill=f_fill, align=center)

    set_c("P1", "2026-27", font=font_header, fill=fill_navy, align=center)

    # Side Table Header
    set_c("R1", "Particulars", font=font_header, fill=fill_navy, align=left)
    set_c("S1", "Till Date", font=font_header, fill=fill_navy, align=center)
    set_c("T1", "Balance", font=font_header, fill=fill_navy, align=center)
    set_c("U1", "Target", font=font_header, fill=fill_navy, align=center)

    cats = [
        ('Edible Oil', 'edible'),
        ('Other liquid', 'other'),
        ('Chemical', 'chemical'),
        ('POL', 'pol'),
    ]

    cur_r = 2
    for cat_lbl, cat_k in cats:
        set_c(f"B{cur_r}", cat_lbl, font=font_bold, fill=fill_white, align=left)
        set_c(f"C{cur_r}", None, fill=fill_white)
        ws.merge_cells(f"B{cur_r}:C{cur_r}")
        for i, m in enumerate(payload['months']):
            col_let = get_column_letter(4 + i)
            val = m['budget'][cat_k]
            set_c(f"{col_let}{cur_r}", val if val > 0 else None, font=font_normal, fmt=NUM_FMT if val > 0 else None)

        cat_summary = payload['summary']['category_summary'][cat_k]
        b_val = cat_summary['budget']
        act_val = cat_summary['actual']
        bal_val = cat_summary['balance']
        set_c(f"P{cur_r}", b_val if b_val > 0 else None, font=font_bold, fmt=NUM_FMT if b_val > 0 else None)

        # Side table
        set_c(f"R{cur_r}", cat_lbl, font=font_bold, align=left)
        set_c(f"S{cur_r}", act_val, font=font_bold, fmt=NUM_FMT)
        set_c(f"T{cur_r}", bal_val if bal_val > 0 else None, font=font_var_neg if bal_val > 0 else font_bold, fmt=NUM_FMT if bal_val > 0 else None)
        set_c(f"U{cur_r}", b_val if b_val > 0 else None, font=font_bold, fmt=NUM_FMT if b_val > 0 else None)
        cur_r += 1

    # Total Cargo Budget
    set_c(f"B{cur_r}", "Total Cargo Budget", font=font_green_bold, fill=fill_soft_green, align=left)
    set_c(f"C{cur_r}", None, fill=fill_soft_green)
    ws.merge_cells(f"B{cur_r}:C{cur_r}")
    for i, m in enumerate(payload['months']):
        col_let = get_column_letter(4 + i)
        set_c(f"{col_let}{cur_r}", m['budget']['total'], font=font_green_bold, fill=fill_soft_green, fmt=NUM_FMT)

    tot_bud = payload['summary']['full_year_budget']
    ytd_act = payload['summary']['ytd_actual']
    tot_bal = payload['summary']['balance']

    set_c(f"P{cur_r}", tot_bud, font=font_green_bold, fill=fill_soft_green, fmt=NUM_FMT)
    set_c(f"R{cur_r}", "Total", font=font_green_bold, fill=fill_soft_green, align=left)
    set_c(f"S{cur_r}", ytd_act, font=font_green_bold, fill=fill_soft_green, fmt=NUM_FMT)
    set_c(f"T{cur_r}", tot_bal if tot_bal > 0 else None, font=font_var_neg, fill=fill_soft_green, fmt=NUM_FMT if tot_bal > 0 else None)
    set_c(f"U{cur_r}", tot_bud, font=font_green_bold, fill=fill_soft_green, fmt=NUM_FMT)
    cur_r += 1

    # Cum Budget
    set_c(f"B{cur_r}", "Cum Budget", font=font_bold, fill=fill_cum_budget, align=left)
    set_c(f"C{cur_r}", None, fill=fill_cum_budget)
    ws.merge_cells(f"B{cur_r}:C{cur_r}")
    for i, m in enumerate(payload['months']):
        col_let = get_column_letter(4 + i)
        set_c(f"{col_let}{cur_r}", m['cum_budget'], font=font_normal, fill=fill_cum_budget, fmt=NUM_FMT)
    set_c(f"P{cur_r}", tot_bud, font=font_bold, fill=fill_cum_budget, fmt=NUM_FMT)
    cur_r += 1

    # Actual
    act_r = cur_r
    set_c(f"B{cur_r}", "Actual", font=font_bold, fill=fill_highlight, align=left)
    set_c(f"C{cur_r}", None, fill=fill_highlight)
    ws.merge_cells(f"B{cur_r}:C{cur_r}")
    for i, m in enumerate(payload['months']):
        col_let = get_column_letter(4 + i)
        val = m['actual']['total'] if m['is_past_or_current'] else None
        set_c(f"{col_let}{cur_r}", val, font=font_bold if val is not None else font_normal, fill=fill_highlight, fmt=NUM_FMT if val is not None else None)
    set_c(f"P{cur_r}", ytd_act, font=font_bold, fill=fill_highlight, fmt=NUM_FMT)
    cur_r += 1

    # % Achieved
    set_c(f"B{cur_r}", "% Achieved", font=font_italic, fill=fill_white, align=left)
    set_c(f"C{cur_r}", None, fill=fill_white)
    ws.merge_cells(f"B{cur_r}:C{cur_r}")
    for i, m in enumerate(payload['months']):
        col_let = get_column_letter(4 + i)
        formula_val = f"={col_let}8/{col_let}6" if m['is_past_or_current'] else None
        set_c(f"{col_let}{cur_r}", formula_val, font=font_italic, fmt="0%" if formula_val else None)
    set_c(f"P{cur_r}", "=P8/P6", font=font_green_bold, fill=fill_white, fmt="0%")
    cur_r += 1

    # Cumulative Achieved
    set_c(f"B{cur_r}", "Cumulative Achieved", font=font_bold, fill=fill_highlight, align=left)
    set_c(f"C{cur_r}", None, fill=fill_highlight)
    ws.merge_cells(f"B{cur_r}:C{cur_r}")
    for i, m in enumerate(payload['months']):
        col_let = get_column_letter(4 + i)
        if not m['is_past_or_current']:
            formula_val = None
        elif i == 0:
            formula_val = f"={col_let}8"
        else:
            prev_let = get_column_letter(4 + i - 1)
            formula_val = f"={prev_let}10+{col_let}8"
        set_c(f"{col_let}{cur_r}", formula_val, font=font_bold if formula_val else font_normal, fill=fill_highlight, fmt=NUM_FMT if formula_val else None)
    set_c(f"P{cur_r}", None, fill=fill_highlight)
    cur_r += 1

    # % Achieved Cumulative
    set_c(f"B{cur_r}", "% Achieved Cumulative", font=font_italic, fill=fill_white, align=left)
    set_c(f"C{cur_r}", None, fill=fill_white)
    ws.merge_cells(f"B{cur_r}:C{cur_r}")
    for i, m in enumerate(payload['months']):
        col_let = get_column_letter(4 + i)
        formula_val = f"={col_let}10/$P$6" if m['is_past_or_current'] else None
        set_c(f"{col_let}{cur_r}", formula_val, font=font_italic, fmt="0%" if formula_val else None)
    set_c(f"P{cur_r}", None, fill=fill_white)
    cur_r += 1

    # Variance
    set_c(f"B{cur_r}", "Variance", font=font_bold, fill=fill_white, align=left)
    set_c(f"C{cur_r}", None, fill=fill_white)
    ws.merge_cells(f"B{cur_r}:C{cur_r}")
    for i, m in enumerate(payload['months']):
        col_let = get_column_letter(4 + i)
        formula_val = f"={col_let}10-{col_let}7" if m['is_past_or_current'] else None
        set_c(f"{col_let}{cur_r}", formula_val, font=font_var_neg if formula_val else font_normal, fmt=NUM_FMT if formula_val else None)
    set_c(f"P{cur_r}", "=P8/P6", font=Font(name=FONT_NAME, size=16, bold=True, color="00B050"), fill=fill_white, align=right, fmt="0.0%")

    # 2. CARGO VOLUMES BAR CHART (Anchored at X1)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Cargo Volumes"
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.legend = None
    chart.width = 16
    chart.height = 7.5
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    data_ref = Reference(ws, min_col=4, min_row=act_r, max_col=15, max_row=act_r)
    cats_ref = Reference(ws, min_col=4, min_row=1, max_col=15, max_row=1)
    chart.add_data(data_ref, from_rows=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "X1")

    # 3. HISTORICAL & QUARTERLY BREAKDOWN TABLE (Under Chart, Rows 12 to 25, Cols X, Y, Z)
    border_b_med = Border(bottom=Side(style="medium", color="B7B7B7"))
    border_t_thin = Border(top=Side(style="thin", color="D9D9D9"))

    hist = payload['summary'].get('historical', {})
    q = payload['summary'].get('quarterly', {})
    fy25_val = hist.get('fy25', 200912.538)
    fy26_val = hist.get('fy26', 1300492.320)
    fy26_q1 = hist.get('fy26_q1', 283589.208)
    fy26_q2 = hist.get('fy26_q2', 455638.864)
    fy26_q3 = hist.get('fy26_q3', 282231.927)
    fy26_q4 = hist.get('fy26_q4', 277032.321)

    q1_27 = q.get('q1', 0.0)
    q2_27 = q.get('q2', 0.0)
    q3_27 = q.get('q3', 0.0)
    q4_27 = q.get('q4', 0.0)
    fy27_val = payload['summary'].get('ytd_actual', 0.0)

    # FY 25
    set_c("X12", "Nov-Mar", font=font_normal, align=left, bdr=None)
    set_c("Y12", fy25_val, font=font_bold, align=right, fmt=NUM_FMT, bdr=None)

    set_c("X13", "FY 25", font=font_bold, align=left, bdr=border_b_med)
    set_c("Y13", fy25_val, font=font_bold, align=right, fmt=NUM_FMT, bdr=border_b_med)

    # FY 26
    set_c("X15", "Q1", font=font_normal, align=left, bdr=None)
    set_c("Y15", fy26_q1, font=font_normal, align=right, fmt=NUM_FMT, bdr=None)
    set_c("Z15", (fy26_q1 / fy26_val) if fy26_val > 0 else None, font=font_normal, align=right, fmt="0%", bdr=None)

    set_c("X16", "Q2", font=font_normal, align=left, bdr=None)
    set_c("Y16", fy26_q2, font=font_normal, align=right, fmt=NUM_FMT, bdr=None)
    set_c("Z16", (fy26_q2 / fy26_val) if fy26_val > 0 else None, font=font_normal, align=right, fmt="0%", bdr=None)

    set_c("X17", "Q3", font=font_normal, align=left, bdr=None)
    set_c("Y17", fy26_q3, font=font_normal, align=right, fmt=NUM_FMT, bdr=None)
    set_c("Z17", (fy26_q3 / fy26_val) if fy26_val > 0 else None, font=font_normal, align=right, fmt="0%", bdr=None)

    set_c("X18", "Q4", font=font_normal, align=left, bdr=None)
    set_c("Y18", fy26_q4, font=font_normal, align=right, fmt=NUM_FMT, bdr=None)
    set_c("Z18", (fy26_q4 / fy26_val) if fy26_val > 0 else None, font=font_normal, align=right, fmt="0%", bdr=None)

    set_c("X19", "FY 26", font=font_bold, align=left, bdr=border_b_med)
    set_c("Y19", fy26_val, font=font_bold, align=right, fmt=NUM_FMT, bdr=border_b_med)

    # FY 27
    set_c("X21", "Q1", font=font_normal, align=left, bdr=None)
    set_c("Y21", q1_27, font=font_normal, align=right, fmt=NUM_FMT, bdr=None)
    set_c("Z21", (q1_27 / fy27_val) if (fy27_val > 0 and q1_27 > 0) else None, font=font_normal, align=right, fmt="0%" if q1_27 > 0 else None, bdr=None)

    set_c("X22", "Q2", font=font_normal, align=left, bdr=None)
    set_c("Y22", q2_27, font=font_normal, align=right, fmt=NUM_FMT, bdr=None)
    set_c("Z22", (q2_27 / fy27_val) if (fy27_val > 0 and q2_27 > 0) else None, font=font_normal, align=right, fmt="0%" if q2_27 > 0 else None, bdr=None)

    set_c("X23", "Q3", font=font_normal, align=left, bdr=None)
    set_c("Y23", q3_27 if q3_27 > 0 else 0.0, font=font_normal, align=right, fmt=NUM_FMT, bdr=None)
    set_c("Z23", "-", font=font_normal, align=right, bdr=None)

    set_c("X24", "Q4", font=font_normal, align=left, bdr=None)
    set_c("Y24", q4_27 if q4_27 > 0 else 0.0, font=font_normal, align=right, fmt=NUM_FMT, bdr=None)
    set_c("Z24", "-", font=font_normal, align=right, bdr=None)

    set_c("X25", "FY 27", font=font_bold, align=left, bdr=border_t_thin)
    set_c("Y25", fy27_val, font=font_bold, align=right, fmt=NUM_FMT, bdr=border_t_thin)

    # 4. TABLE GAP SPACING (Rows before Bottom Vessel Table)
    cur_r = max(cur_r + 3, 27)

    # 5. BOTTOM VESSEL TABLE
    # Floating category summary row
    tot_edible = sum(v['edible'] for v in payload['vessels'])
    tot_other = sum(v['other'] for v in payload['vessels'])
    tot_chem = sum(v['chemical'] for v in payload['vessels'])
    tot_pol = sum(v['pol'] for v in payload['vessels'])

    set_c(f"H{cur_r}", tot_edible, font=font_floating_sum, fill=fill_white, bdr=None)
    set_c(f"I{cur_r}", tot_other, font=font_floating_sum, fill=fill_white, bdr=None)
    set_c(f"J{cur_r}", tot_chem, font=font_floating_sum, fill=fill_white, bdr=Border())
    set_c(f"K{cur_r}", tot_pol, font=font_floating_sum, fill=fill_white, bdr=Border())
    cur_r += 1

    # Vessel Table Headers
    v_hdrs = [
        ("A", "#", center), ("B", "Month", center), ("C", "Vessel Name", left),
        ("D", "Cargo", left), ("E", "Customer", left), ("F", "Quantity", right),
        ("G", "Month Total", right),
        ("H", "Edible Oil", right), ("I", "Other liquid", right), ("J", "Chemical", right),
        ("K", "POL", right)
    ]
    for col_l, h_title, al in v_hdrs:
        set_c(f"{col_l}{cur_r}", h_title, font=font_header, fill=fill_sub_blue, align=al)

    # Determine last vessel index for each month
    month_totals = {}
    last_vessel_idx_for_month = {}
    for idx, v in enumerate(payload['vessels']):
        m_lbl = v['month']
        month_totals[m_lbl] = month_totals.get(m_lbl, 0.0) + (v['quantity'] or 0.0)
        last_vessel_idx_for_month[m_lbl] = idx

    # Alternating month shading
    last_month = ''
    month_flip = False

    for idx, v in enumerate(payload['vessels']):
        cur_r += 1
        if v['month'] != last_month:
            month_flip = not month_flip
            last_month = v['month']
        row_fill = fill_month_alt if month_flip else fill_white

        is_last = (last_vessel_idx_for_month.get(v['month']) == idx)
        m_tot_val = month_totals.get(v['month']) if is_last else None

        set_c(f"A{cur_r}", v['sr_no'], font=font_normal, fill=row_fill, align=center)
        set_c(f"B{cur_r}", v['month'], font=font_bold, fill=row_fill, align=center)
        set_c(f"C{cur_r}", v['vessel_name'], font=font_bold, fill=row_fill, align=left)
        set_c(f"D{cur_r}", v['cargo'], font=font_normal, fill=row_fill, align=left)
        set_c(f"E{cur_r}", v['customer'], font=font_normal, fill=row_fill, align=left)
        set_c(f"F{cur_r}", v['quantity'], font=font_bold, fill=row_fill, fmt=NUM_FMT)

        if is_last:
            set_c(f"G{cur_r}", m_tot_val, font=font_bold, fill=fill_amber, fmt=NUM_FMT)
        else:
            set_c(f"G{cur_r}", None, fill=row_fill)

        set_c(f"H{cur_r}", v['edible'] if v['edible'] > 0 else None, font=font_normal, fill=row_fill, fmt=NUM_FMT)
        set_c(f"I{cur_r}", v['other'] if v['other'] > 0 else None, font=font_normal, fill=row_fill, fmt=NUM_FMT)
        set_c(f"J{cur_r}", v['chemical'] if v['chemical'] > 0 else None, font=font_normal, fill=row_fill, fmt=NUM_FMT)
        set_c(f"K{cur_r}", v['pol'] if v['pol'] > 0 else None, font=font_normal, fill=row_fill, fmt=NUM_FMT)

    # Total row for vessels
    cur_r += 1
    set_c(f"A{cur_r}", f"Total ({len(payload['vessels'])} vessel calls)", font=font_green_bold, fill=fill_soft_green, align=left, bdr=border_total)
    for col_l in ('B', 'C', 'D', 'E'):
        set_c(f"{col_l}{cur_r}", None, fill=fill_soft_green, bdr=border_total)
    ws.merge_cells(f"A{cur_r}:E{cur_r}")

    tot_qty = sum(v['quantity'] for v in payload['vessels'])
    set_c(f"F{cur_r}", tot_qty, font=font_green_bold, fill=fill_soft_green, fmt=NUM_FMT, bdr=border_total)
    set_c(f"G{cur_r}", None, fill=fill_soft_green, bdr=border_total)
    set_c(f"H{cur_r}", tot_edible, font=font_green_bold, fill=fill_soft_green, fmt=NUM_FMT, bdr=border_total)
    set_c(f"I{cur_r}", tot_other, font=font_green_bold, fill=fill_soft_green, fmt=NUM_FMT, bdr=border_total)
    set_c(f"J{cur_r}", tot_chem, font=font_green_bold, fill=fill_soft_green, fmt=NUM_FMT, bdr=border_total)
    set_c(f"K{cur_r}", tot_pol, font=font_green_bold, fill=fill_soft_green, fmt=NUM_FMT, bdr=border_total)

    # Column Widths
    col_widths = {
        'A': 6, 'B': 12, 'C': 30, 'D': 20, 'E': 32, 'F': 16, 'G': 16,
        'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 14, 'M': 14, 'N': 14,
        'O': 14, 'P': 16, 'Q': 4, 'R': 18, 'S': 16, 'T': 16, 'U': 16,
        'V': 4, 'W': 4, 'X': 14, 'Y': 16, 'Z': 10
    }
    for col_l, width in col_widths.items():
        ws.column_dimensions[col_l].width = width

    ws.sheet_view.showGridLines = True
    return ws


@bp.route('/api/module/RP01/dpr/export-bvsa')
@login_required
def dpr_export_bvsa():
    """Export the combined 2-sheet workbook (Sheet 1: DPR, Sheet 2: BVsA FY 2027)."""
    return dpr_export()



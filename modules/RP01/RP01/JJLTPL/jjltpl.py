from .. import bp

from flask import (
    render_template,
    session,
    redirect,
    url_for,
    request,
    jsonify,
    Response,
    send_file,
)

from functools import wraps
from datetime import datetime, time, timedelta
from io import BytesIO
from database import get_db, get_cursor, get_user_permissions

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


MODULE_CODE = 'RP01'

# Flip this to True temporarily if numbers still look wrong — it prints
# per parcel-op target vs. actual to stderr so you can see exactly what
# the aggregation is doing at runtime.
_JJLTPL_DEBUG = False


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def get_perms():
    if session.get('is_admin'):
        return {'can_read': 1, 'can_add': 1, 'can_edit': 1, 'can_delete': 1}
    return get_user_permissions(session.get('user_id'), MODULE_CODE)


TERMINAL_BERTH_MAP = {
    'JJLTPL': ['LB-03', 'LB-04'],
}
DEFAULT_TERMINAL = 'JJLTPL'

MEDIUM_DRY_BULK = 'Dry Bulk'
MEDIUM_BREAK_BULK = 'Break Bulk'
MEDIUM_LIQUID_BULK = 'Liquid Bulk'


def _jjltpl_parse_date(date_str):
    if date_str:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    return datetime.now().date()


def _jjltpl_window(selected_date):
    window_end = datetime.combine(selected_date, time(7, 0, 0))
    window_start = window_end - timedelta(days=1)
    return window_start, window_end


def _jjltpl_month_window(selected_date, window_end=None):
    d = selected_date.date() if isinstance(selected_date, datetime) else selected_date
    if d.day == 1:
        prev_month = 12 if d.month == 1 else d.month - 1
        prev_year = d.year - 1 if d.month == 1 else d.year
        month_start = datetime(prev_year, prev_month, 1, 7, 0, 0)
    else:
        month_start = datetime.combine(d.replace(day=1), time(7, 0, 0))
    month_end = datetime.combine(d, time(7, 0, 0))
    return month_start, month_end


def _jjltpl_year_window(selected_date, window_end):

    if selected_date.month >= 4:
        fy_start = datetime(selected_date.year, 4, 1, 7, 0, 0)
    else:
        fy_start = datetime(selected_date.year - 1, 4, 1, 7, 0, 0)

    return fy_start, window_end


def _jjltpl_fin_year_label(selected_date):
    """
    Financial-year label matching the `fin_year` column in
    mis_vessel_master, e.g. 2026-04-.. through 2027-03-.. -> "2026-27".

    NOTE: adjust the format below if your DB actually stores it as
    "2026-2027", "FY26-27", etc.
    """
    if selected_date.month >= 4:
        start_year = selected_date.year
    else:
        start_year = selected_date.year - 1
    end_year_short = (start_year + 1) % 100
    return f"{start_year}-{end_year_short:02d}"


def _lueu_parse_ids(csv):
    return [int(x) for x in str(csv or '').split(',') if str(x).strip().isdigit()]


def _lueu_hours(f, t):
    """Duration in hours between two 'HH:MM' strings (wraps past midnight).
    Mirrors LUEU01/model.py::_hours exactly."""
    try:
        fh, fm = (int(x) for x in str(f).split(':')[:2])
        th, tm = (int(x) for x in str(t).split(':')[:2])
    except (ValueError, AttributeError):
        return 0.0
    mins = (th * 60 + tm) - (fh * 60 + fm)
    if mins < 0:
        mins += 1440
    return mins / 60.0


def _lueu_target_qty(cur, parcel_ids_csv, op_qty, operation_type):
    """Current target quantity for a parcel-op: sum of its live VCN parcel
    quantities (import or export source table), falling back to the
    parcel-op's own snapshot quantity. Mirrors model.py::_single_parcel_target
    / the target-resolution block in get_started_parcels."""
    ids = _lueu_parse_ids(parcel_ids_csv)
    tbl = 'vcn_export_cargo_declaration' if operation_type == 'Export' else 'vcn_consigners'
    total = 0.0
    if ids:
        cur.execute(f'SELECT quantity FROM {tbl} WHERE id = ANY(%s)', [ids])
        for r in cur.fetchall():
            try:
                total += float(str(r['quantity']).replace(',', '')) if r['quantity'] else 0.0
            except (ValueError, TypeError):
                pass
    return total or float(op_qty or 0)


def _lueu_is_shortclose_row(r):
    """
    Detect a shortclose row from the `remarks` text column (e.g. operators
    typing "Shortclose", "Short Close", "SHORT-CLOSE" etc.) rather than
    relying solely on the is_shortclose boolean flag, since that flag is
    not reliably set (confirmed: rows with remarks="Short Close" can still
    have is_shortclose=false in the data). Falls back to is_shortclose if
    the remarks text doesn't match, in case some rows only set the flag.
    """
    remarks = str(r.get('remarks') or '').strip().lower()
    if 'short' in remarks and 'close' in remarks:
        return True
    return bool(r.get('is_shortclose'))


def _lueu_log_aggregate(cur, parcel_op_id, target):
    """
    Aggregation of lueu_parcel_log rows for one parcel-op:
      - real_qty = sum of qty on normal (non-shortclose) rows, capped so
        it never exceeds the effective target.
      - shortclose_qty = sum of qty on rows whose remark marks them as a
        shortclose. This REDUCES the target (vessel closed short of the
        full BL) rather than counting as delivered cargo — so it is NOT
        added to real_qty and does not count toward hours/avg_rate.
      - effective_target = target - shortclose_qty (never below 0).

    Returns (real_qty, hours, shortclose_qty, effective_target).
    """
    cur.execute('''SELECT from_time, to_time, COALESCE(quantity,0) AS q,
                          is_shortclose, remarks
                   FROM lueu_parcel_log
                   WHERE parcel_op_id=%s AND is_deleted IS NOT TRUE
                   ORDER BY entry_date, from_time NULLS LAST, id''', [parcel_op_id])

    rows = cur.fetchall()

    # First pass: total shortclose qty reduces the target.
    shortclose_qty = sum(
        float(r['q'] or 0) for r in rows if _lueu_is_shortclose_row(r)
    )
    effective_target = max(target - shortclose_qty, 0) if target > 0 else target

    # Second pass: sum real (non-shortclose) rows, capped at effective_target.
    real_qty, hours = 0.0, 0.0
    for r in rows:
        if _lueu_is_shortclose_row(r):
            continue
        if effective_target > 0 and real_qty >= effective_target - 1e-6:
            break
        real_qty += float(r['q'] or 0)
        hours += _lueu_hours(r['from_time'], r['to_time'])

    return real_qty, hours, shortclose_qty, effective_target


def _jjltpl_actual_qty_for_rows(cur, rows, label=""):
    """
    Shared helper: given raw rows containing parcel_op_id / parcel_ids /
    op_qty / operation_type, resolve each parcel-op's TARGET (BL) quantity,
    then aggregate the lueu_parcel_log to find what has ACTUALLY been
    logged so far (real + shortclose, capped at target). Sums that across
    all distinct parcel-ops.

    Deduplicates on parcel_op_id in case the SQL join produces the same
    parcel-op more than once (e.g. multiple ldud_header rows per vessel),
    which would otherwise double-count it.

    NOTE: this is a CUMULATIVE (lifetime-to-date) figure. It is used for
    MONTH/YEAR, where the underlying query already scopes vessels by
    cast_off_datetime falling inside the period, so "actual to date" is
    effectively "final total for that call". DAY does NOT use this —
    see _jjltpl_bulk_tons, which sums log entries directly against the
    24hr window instead.
    """
    qty = 0.0
    seen_ids = set()

    for r in rows:
        pid = r["parcel_op_id"]
        if not pid or pid in seen_ids:
            continue
        seen_ids.add(pid)

        target = _lueu_target_qty(cur, r["parcel_ids"], r["op_qty"], r["operation_type"])
        real_qty, hours, shortclose_qty, effective_target = _lueu_log_aggregate(cur, pid, target)
        # Actual quantity handled = what was really pumped/loaded, i.e.
        # real_qty only. Shortclose qty is NOT delivered cargo — it's a
        # reduction of the target — so it's excluded from the tons total.
        actual = real_qty
        qty += actual

        if _JJLTPL_DEBUG:
            import sys
            print(
                f"[JJLTPL DEBUG {label}] parcel_op_id={pid} "
                f"op_qty={r['op_qty']} target={target} "
                f"shortclose_qty={shortclose_qty} effective_target={effective_target} "
                f"real_qty={real_qty} remaining={max(effective_target - real_qty, 0)}",
                file=sys.stderr,
            )

    return qty


def _parse_entry_date(entry_date):
    """
    entry_date may come back from the DB as a real date/datetime object,
    or as plain text (this schema stores several date/time columns as
    text elsewhere, cast with ::timestamp in SQL). Handle both so a
    string value doesn't raise inside datetime.combine() and blow up
    the request with a 500.
    """
    if entry_date is None:
        return None
    if isinstance(entry_date, datetime):
        return entry_date.date()
    if hasattr(entry_date, 'year') and hasattr(entry_date, 'month'):
        # already a date object
        return entry_date
    s = str(entry_date).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Last resort: try fromisoformat (handles date or datetime strings)
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return None


def _jjltpl_bulk_tons(cur, period_start, period_end, berths):
    """
    DAY quantity & vessel count:
    Sums log entries (lueu_parcel_log) falling strictly within [period_start, period_end)
    for vessels at these berths.
    Shows only the selected 24-hr operational day period data, excluding running
    or cumulative multi-day totals from previous days.
    Ignores is_deleted rows and respects short-close reductions.
    """
    cur.execute("""
        SELECT
            log.entry_date,
            log.from_time,
            log.to_time,
            COALESCE(log.quantity, 0) AS q,
            log.is_shortclose,
            log.remarks,
            vh.id AS vh_id
        FROM lueu_parcel_log log
        JOIN ldud_parcel_ops po ON po.id = log.parcel_op_id
        JOIN ldud_header lh ON lh.id = po.ldud_id
        JOIN vcn_header vh ON vh.id = lh.vcn_id
        WHERE vh.berth_name = ANY(%s)
          AND log.is_deleted IS NOT TRUE
    """, (berths,))

    rows = cur.fetchall()
    log_qty = 0.0
    vessel_ids = set()

    for r in rows:
        if _lueu_is_shortclose_row(r):
            continue

        entry_date = _parse_entry_date(r['entry_date'])
        from_time = r['from_time']
        if not entry_date or not from_time:
            continue
        try:
            fh, fm = (int(x) for x in str(from_time).split(':')[:2])
        except (ValueError, AttributeError):
            continue

        entry_dt = datetime.combine(entry_date, time(fh, fm))
        if period_start <= entry_dt < period_end:
            log_qty += float(r['q'] or 0)
            vessel_ids.add(r['vh_id'])

    return {
        MEDIUM_DRY_BULK: 0.0,
        MEDIUM_BREAK_BULK: 0.0,
        MEDIUM_LIQUID_BULK: log_qty,
        "bulk_total": log_qty,
        "vessel_count": len(vessel_ids),
    }


def _jjltpl_fy_bulk_tons(cur, fin_year):

    cur.execute("""
        SELECT
            COALESCE(SUM(quantity),0) AS qty
        FROM mis_vessel_master
        WHERE fin_year = %s
    """, (fin_year,))

    qty = float(cur.fetchone()["qty"] or 0)

    return {
        MEDIUM_DRY_BULK: 0.0,
        MEDIUM_BREAK_BULK: 0.0,
        MEDIUM_LIQUID_BULK: qty,
        "bulk_total": qty,
    }


def _jjltpl_fy_bulk_vessel_count(cur, fin_year):
    """
    Count of vessels in mis_vessel_master for the given financial year.
    """
    cur.execute("""
        SELECT COUNT(*) AS cnt
        FROM mis_vessel_master
        WHERE fin_year = %s
    """, (fin_year,))

    row = cur.fetchone()
    return row["cnt"] if row and row["cnt"] else 0


def _jjltpl_vessels_on_berth(cur, window_start, window_end, berths):

    cur.execute("""
        SELECT
            vh.berth_name AS berth,
            vh.via_number AS via,
            vh.vessel_name,
            vh.operation_type,

            po.id AS parcel_op_id,
            po.parcel_ids,
            COALESCE(
                NULLIF(TRIM(vc.cargo_code), ''),
                NULLIF(TRIM(REGEXP_REPLACE(po.cargo_name, '\\s*\\[.*\\]', '')), ''),
                NULLIF(TRIM(vc2.cargo_code), ''),
                NULLIF(TRIM(REGEXP_REPLACE(vh.cargo_type, '\\s*\\[.*\\]', '')), ''),
                TRIM(po.cargo_name),
                TRIM(vh.cargo_type)
            ) AS cargo_type,
            po.quantity AS op_qty,

            NULLIF(lh.alongside_datetime,'')::timestamp AS alongside_datetime,

            NULLIF(po.start_dt,'')::timestamp AS start_dt,
            NULLIF(po.expected_start,'')::timestamp AS expected_start,

            COALESCE(po.expected_flow_rate,0) AS expected_flow_rate

        FROM vcn_header vh

        JOIN ldud_header lh
            ON lh.vcn_id = vh.id

        LEFT JOIN ldud_parcel_ops po
            ON po.ldud_id = lh.id

        LEFT JOIN vessel_cargo vc ON (
            UPPER(TRIM(vc.cargo_name)) = UPPER(TRIM(po.cargo_name))
            OR UPPER(TRIM(vc.cargo_code)) = UPPER(TRIM(po.cargo_name))
            OR UPPER(TRIM(vc.cargo_name)) = UPPER(TRIM(REGEXP_REPLACE(po.cargo_name, '\\s*\\[.*\\]', '')))
            OR UPPER(TRIM(vc.cargo_code)) = UPPER(TRIM(REGEXP_REPLACE(po.cargo_name, '\\s*\\[.*\\]', '')))
        )

        LEFT JOIN vessel_cargo vc2 ON (
            UPPER(TRIM(vc2.cargo_name)) = UPPER(TRIM(vh.cargo_type))
            OR UPPER(TRIM(vc2.cargo_code)) = UPPER(TRIM(vh.cargo_type))
            OR UPPER(TRIM(vc2.cargo_name)) = UPPER(TRIM(REGEXP_REPLACE(vh.cargo_type, '\\s*\\[.*\\]', '')))
            OR UPPER(TRIM(vc2.cargo_code)) = UPPER(TRIM(REGEXP_REPLACE(vh.cargo_type, '\\s*\\[.*\\]', '')))
        )

        WHERE vh.berth_name = ANY(%s)

        AND NULLIF(lh.alongside_datetime,'') IS NOT NULL
        AND NULLIF(lh.alongside_datetime,'')::timestamp <= %s

        AND (
            -- Still on berth
            NULLIF(lh.cast_off_datetime,'') IS NULL

            -- Or cast off happens after this report window
            OR NULLIF(lh.cast_off_datetime,'')::timestamp > %s
        )

        ORDER BY
            vh.berth_name, po.id
    """, (berths, window_end, window_end))

    raw_rows = cur.fetchall()

    rows = []

    for r in raw_rows:

        expected_completion = None

        if r["parcel_op_id"]:

            # Same target resolution + log aggregation as the rest of the
            # report. shortclose qty reduces the effective target rather
            # than counting as delivered.
            target = _lueu_target_qty(cur, r["parcel_ids"], r["op_qty"], r["operation_type"])
            real_qty, hours, shortclose_qty, effective_target = _lueu_log_aggregate(
                cur, r["parcel_op_id"], target
            )
            remaining_qty = max(effective_target - real_qty, 0)

            # Matches lueu01.html's etcText() exactly:
            #   ETC = start_dt + (effective_target / avg_rate) hours
            # — projected from the fixed start time using the FULL
            # (shortclose-adjusted) target, not "remaining" or "now".
            # Only shown while remaining > 0.
            if remaining_qty > 0:

                if r["start_dt"] and hours > 0 and real_qty > 0:
                    avg_rate = real_qty / hours
                    if avg_rate > 0 and effective_target > 0:
                        expected_completion = r["start_dt"] + timedelta(
                            hours=(effective_target / avg_rate)
                        )

                elif (
                    r["expected_start"]
                    and r["expected_flow_rate"]
                    and float(r["expected_flow_rate"]) > 0
                    and effective_target > 0
                ):
                    # Mirrors the separate "ETC (Exp)" chip: expected_start +
                    # (effective_target / expected_flow_rate), used only when
                    # there's no actual start yet.
                    expected_completion = r["expected_start"] + timedelta(
                        hours=(effective_target / float(r["expected_flow_rate"]))
                    )

        rows.append({
            "berth": r["berth"],
            "via": r["via"],
            "vessel_name": r["vessel_name"],
            "cargo": r["cargo_type"],

            "alongside_datetime":
                r["alongside_datetime"].isoformat()
                if r["alongside_datetime"] else None,

            "expected_completion":
                expected_completion.isoformat()
                if expected_completion else None,

            "anchor_reason": None
        })

    # ------------------------------------------------------------------
    # Collapse multiple parcel/cargo rows belonging to the same vessel
    # call into a single row:
    #   - cargoes joined as "cargo1/cargo2" (dedup, order preserved)
    #   - expected completion = the LATEST (max) ETC among that
    #     vessel's parcels, since the vessel isn't done until all its
    #     parcels are done
    # Grouped by (berth, via, vessel_name, alongside_datetime) — these
    # four fields are identical across all parcel-ops of one vessel
    # call, so this is a safe stable key without needing a header id.
    # ------------------------------------------------------------------
    grouped = {}
    order = []
    for row in rows:
        key = (row["berth"], row["via"], row["vessel_name"], row["alongside_datetime"])
        if key not in grouped:
            grouped[key] = {
                "berth": row["berth"],
                "via": row["via"],
                "vessel_name": row["vessel_name"],
                "cargo_list": [],
                "alongside_datetime": row["alongside_datetime"],
                "expected_completion_list": [],
                "anchor_reason": row["anchor_reason"],
            }
            order.append(key)
        g = grouped[key]
        if row["cargo"] and row["cargo"] not in g["cargo_list"]:
            g["cargo_list"].append(row["cargo"])
        if row["expected_completion"]:
            g["expected_completion_list"].append(row["expected_completion"])

    rows = []
    for key in order:
        g = grouped[key]
        # ISO-8601 strings sort chronologically, so max() gives the
        # latest ETC without needing to parse back to datetime.
        latest_etc = max(g["expected_completion_list"]) if g["expected_completion_list"] else None
        rows.append({
            "berth": g["berth"],
            "via": g["via"],
            "vessel_name": g["vessel_name"],
            "cargo": "/".join(g["cargo_list"]) if g["cargo_list"] else None,
            "alongside_datetime": g["alongside_datetime"],
            "expected_completion": latest_etc,
            "anchor_reason": g["anchor_reason"],
        })

    # ------------------------------------------------------------------
    # Always show all berths (LB-03 and LB-04)
    # ------------------------------------------------------------------
    existing_berths = {row["berth"] for row in rows}

    for berth in berths:
        if berth not in existing_berths:
            rows.append({
                "berth": berth,
                "via": None,
                "vessel_name": None,
                "cargo": None,
                "alongside_datetime": None,
                "expected_completion": None,
                "anchor_reason": None
            })

    # Keep berth order as defined in TERMINAL_BERTH_MAP
    berth_order = {b: i for i, b in enumerate(berths)}
    rows.sort(key=lambda x: berth_order.get(x["berth"], 999))

    return rows


def _jjltpl_month_bulk_tons(cur, period_start, period_end, berths):
    """
    MONTH quantity, based on vessels whose cast_off_datetime (in
    ldud_header) falls within the period — valued at direct SUM(log.quantity).
    """
    cur.execute("""
        SELECT
            COALESCE(SUM(log.quantity), 0) AS qty
        FROM ldud_header lh
        JOIN vcn_header vh
            ON vh.id = lh.vcn_id
        JOIN ldud_parcel_ops po
            ON po.ldud_id = lh.id
        JOIN lueu_parcel_log log
            ON log.parcel_op_id = po.id
        WHERE vh.berth_name = ANY(%s)
          AND log.is_deleted IS NOT TRUE
          AND log.is_shortclose IS NOT TRUE
          AND LOWER(COALESCE(log.remarks, '')) NOT LIKE '%%short%%close%%'
          AND NULLIF(lh.cast_off_datetime, '') IS NOT NULL
          AND NULLIF(lh.cast_off_datetime, '')::timestamp >= %s
          AND NULLIF(lh.cast_off_datetime, '')::timestamp < %s
    """, (berths, period_start, period_end))

    row = cur.fetchone()
    qty = float(row["qty"] or 0)

    return {
        MEDIUM_DRY_BULK: 0.0,
        MEDIUM_BREAK_BULK: 0.0,
        MEDIUM_LIQUID_BULK: qty,
        "bulk_total": qty,
    }


def _jjltpl_bulk_vessel_count(cur, period_start, period_end, berths):
    """
    DAY vessel count: vessels whose cast_off_datetime (cast off complete)
    falls strictly within [period_start, period_end).
    """
    cur.execute("""
        SELECT COUNT(DISTINCT vh.id) AS cnt
        FROM ldud_header lh
        JOIN vcn_header vh
            ON vh.id = lh.vcn_id
        WHERE vh.berth_name = ANY(%s)
          AND NULLIF(lh.cast_off_datetime, '') IS NOT NULL
          AND NULLIF(lh.cast_off_datetime, '')::timestamp >= %s
          AND NULLIF(lh.cast_off_datetime, '')::timestamp < %s
    """, (berths, period_start, period_end))

    row = cur.fetchone()
    return row["cnt"] if row and row["cnt"] else 0


def _jjltpl_month_bulk_vessel_count(cur, period_start, period_end, berths):
    """
    MONTH/YEAR-month-addition vessel count: vessels whose cast_off_datetime
    falls WITHIN [period_start, period_end) — matches the same definition
    _jjltpl_month_bulk_tons uses for tons, so the two figures agree.
    (_jjltpl_bulk_vessel_count is deliberately NOT reused here — it counts
    vessels still sitting on berth right now, which is a different set.)
    """
    cur.execute("""
        SELECT COUNT(DISTINCT vh.id) AS cnt
        FROM ldud_header lh
        JOIN vcn_header vh
            ON vh.id = lh.vcn_id
        WHERE vh.berth_name = ANY(%s)
          AND NULLIF(lh.cast_off_datetime, '') IS NOT NULL
          AND NULLIF(lh.cast_off_datetime, '')::timestamp >= %s
          AND NULLIF(lh.cast_off_datetime, '')::timestamp < %s
    """, (berths, period_start, period_end))

    row = cur.fetchone()
    return row["cnt"] if row and row["cnt"] else 0


def _jjltpl_period_row(cur, label, period_start, period_end, terminal, berths, fin_year=None):

    if label == "YEAR":
        # Existing financial year values
        tons = _jjltpl_fy_bulk_tons(cur, fin_year)
        vessel_count = _jjltpl_fy_bulk_vessel_count(cur, fin_year)

        # Current month values (cast-off-based, matching month_tons logic)
        month_tons = _jjltpl_month_bulk_tons(cur, period_start, period_end, berths)
        month_vessels = _jjltpl_month_bulk_vessel_count(
            cur,
            period_start,
            period_end,
            berths
        )

        # Add month values to existing year values
        tons[MEDIUM_DRY_BULK] += month_tons[MEDIUM_DRY_BULK]
        tons[MEDIUM_BREAK_BULK] += month_tons[MEDIUM_BREAK_BULK]
        tons[MEDIUM_LIQUID_BULK] += month_tons[MEDIUM_LIQUID_BULK]
        tons["bulk_total"] += month_tons["bulk_total"]

        vessel_count += month_vessels
    elif label == "MONTH":
        tons = _jjltpl_month_bulk_tons(cur, period_start, period_end, berths)
        vessel_count = _jjltpl_month_bulk_vessel_count(
            cur,
            period_start,
            period_end,
            berths
        )
    else:
        tons = _jjltpl_bulk_tons(
            cur,
            period_start,
            period_end,
            berths
        )
        vessel_count = _jjltpl_bulk_vessel_count(
            cur,
            period_start,
            period_end,
            berths
        )

    return {
        "period": label,
        "bulk_vessels": vessel_count,
        "dry_bulk_tons": round(tons[MEDIUM_DRY_BULK], 3),
        "break_bulk_tons": round(tons[MEDIUM_BREAK_BULK], 3),
        "liquid_bulk_tons": round(tons[MEDIUM_LIQUID_BULK], 3),
        "bulk_total_tons": round(tons["bulk_total"], 3),
    }

def _jjltpl_report_payload(selected_date, terminal):
    berths = TERMINAL_BERTH_MAP.get(terminal, [])
    window_start, window_end = _jjltpl_window(selected_date)
    month_start, _ = _jjltpl_month_window(selected_date, window_end)
    year_start, _ = _jjltpl_year_window(selected_date, window_end)
    fin_year = _jjltpl_fin_year_label(selected_date)

    conn = get_db()
    try:
        cur = get_cursor(conn)
        vessels_on_berth = _jjltpl_vessels_on_berth(cur, window_start, window_end, berths)
        traffic_rows = [
            _jjltpl_period_row(cur, 'DAY', window_start, window_end, terminal, berths),
            _jjltpl_period_row(cur, 'MONTH', month_start, window_end, terminal, berths),
            _jjltpl_period_row(cur, 'YEAR', year_start, window_end, terminal, berths, fin_year=fin_year),
        ]
    finally:
        conn.close()

    return {
        'terminal': terminal,
        'date': selected_date.strftime('%Y-%m-%d'),
        'date_display': selected_date.strftime('%d-%m-%Y'),
        'window_start': window_start.isoformat(),
        'window_end': window_end.isoformat(),
        'vessels_on_berth': vessels_on_berth,
        'traffic_rows': traffic_rows,
    }


@bp.route('/module/RP01/jjltpl/')
@login_required
def jjltpl_page():
    return render_template("jjltpl.html")


@bp.route('/api/module/RP01/jjltpl/data')
@login_required
def jjltpl_data():
    selected_date = _jjltpl_parse_date(request.args.get('date'))
    terminal = request.args.get('terminal', DEFAULT_TERMINAL)
    try:
        return jsonify(_jjltpl_report_payload(selected_date, terminal))
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(tb)  # still goes to server console/log as usual
        return jsonify({
            'error': str(e),
            'error_type': type(e).__name__,
            'traceback': tb,
        }), 500


# ---------------------------------------------------------------------------
# EXCEL EXPORT — same layout as the reference report, built from the exact
# same payload the UI table uses (_jjltpl_report_payload), so every number
# here is the real DB value, not a placeholder.
# ---------------------------------------------------------------------------

def _fmt_dt_display(iso):
    if not iso:
        return None
    try:
        d = datetime.fromisoformat(iso)
    except ValueError:
        return iso
    return d.strftime('%d-%m-%Y %H:%M')


def _jjltpl_build_workbook(payload):
    wb = Workbook()
    ws = wb.active
    ws.title = payload['terminal']

    FONT_NAME = "Arial"
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_header = PatternFill("solid", fgColor="BCD6EE")
    fill_section = PatternFill("solid", fgColor="DDEBF7")
    fill_month = PatternFill("solid", fgColor="FFF2A8")
    fill_total = PatternFill("solid", fgColor="FCE0CD")
    fill_white = PatternFill("solid", fgColor="FFFFFF")

    font_header = Font(name=FONT_NAME, bold=True, size=10)
    font_section = Font(name=FONT_NAME, bold=True, size=10)
    font_normal = Font(name=FONT_NAME, size=10)
    font_value = Font(name=FONT_NAME, size=10, color="1F4E78")
    font_total = Font(name=FONT_NAME, bold=True, size=10)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    NUM_FMT = '#,##0.000;(#,##0.000);"-"'
    INT_FMT = '#,##0;(#,##0);"-"'

    def set_cell(coord, value, font=font_normal, fill=None, align=center, fmt=None):
        c = ws[coord]
        c.value = value
        c.font = font
        c.border = border
        c.alignment = align
        if fill:
            c.fill = fill
        if fmt:
            c.number_format = fmt
        return c

    def merge(rng, value=None, font=font_normal, fill=None, align=center, fmt=None):
        ws.merge_cells(rng)
        top_left = rng.split(":")[0]
        set_cell(top_left, value, font=font, fill=fill, align=align, fmt=fmt)
        for row in ws[rng]:
            for cell in row:
                cell.border = border
                if fill:
                    cell.fill = fill

    widths = {"A": 20, "B": 14, "C": 14, "D": 22, "E": 12, "F": 20, "G": 22,
              "H": 24, "I": 14, "J": 14, "K": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---- Terminal summary strip ----
    set_cell("A1", "TERMINAL", font=font_section, fill=fill_section, align=left)
    set_cell("B1", payload['terminal'], font=font_value, fill=fill_white)
    merge("C1:D1", "Upto Previous Month TEUs", font=font_header, fill=fill_header)
    merge("E1:F1", "Upto Previous Month TONs", font=font_header, fill=fill_header)
    set_cell("G1", "Date", font=font_header, fill=fill_header)
    set_cell("H1", payload['date_display'], font=font_value, fill=fill_white)

    set_cell("A2", "", fill=fill_section)
    set_cell("B2", "", fill=fill_white)
    merge("C2:D2", None, fill=fill_white)
    merge("E2:F2", None, fill=fill_white)
    merge("G2:H2", None, fill=fill_white)

    # ---- Vessels on Berth (real data) ----
    vessels = payload['vessels_on_berth']
    r = 4
    n_vessel_rows = max(len(vessels), 1)
    merge(f"A{r}:A{r + n_vessel_rows - 1}", "VESSELS ON BERTH",
          font=font_section, fill=fill_section, align=left)

    headers = ["Berth", "Via", "Vessel Name", "Cargo", "Alongside (Date/Time)",
               "Expected Completion (Date/Time)", "Reason if Vessel Anchored before Berthing"]
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        set_cell(f"{col}{r}", h, font=font_header, fill=fill_header)

    if vessels:
        for j, v in enumerate(vessels):
            rr = r + 1 + j
            vals = [
                v.get('berth'), v.get('via'), v.get('vessel_name'), v.get('cargo'),
                _fmt_dt_display(v.get('alongside_datetime')),
                _fmt_dt_display(v.get('expected_completion')),
                v.get('anchor_reason'),
            ]
            for i, val in enumerate(vals):
                col = get_column_letter(2 + i)
                set_cell(f"{col}{rr}", val, font=font_value, fill=fill_white,
                          align=left if i in (2, 6) else center)
        last_vessel_row = r + len(vessels)
    else:
        rr = r + 1
        merge(f"B{rr}:H{rr}", "No vessels on berth", font=font_normal, fill=fill_white)
        last_vessel_row = rr

    # ---- Traffic Throughput — TEU columns (no data source: left blank) +
    #      Tons columns (real data) ----
    traffic = payload['traffic_rows']
    r2 = last_vessel_row + 2
    merge(f"A{r2}:A{r2 + len(traffic) - 1}", "TRAFFIC THROUGHPUT (TEUS)",
          font=font_section, fill=fill_section, align=left)

    tt_headers = ["Period", "Container Vessels", "Imp TEUs", "Exp TEUs", "Total TEUs",
                  "Bulk Vessels", "Dry Bulk Tons", "Break Bulk Tons", "Liquid Bulk Tons",
                  "Bulk Total Tons"]
    for i, h in enumerate(tt_headers):
        col = get_column_letter(2 + i)
        set_cell(f"{col}{r2}", h, font=font_header, fill=fill_header)

    for k, row_data in enumerate(traffic):
        rr = r2 + 1 + k
        is_month = row_data['period'] == 'MONTH'
        row_fill = fill_month if is_month else fill_white
        set_cell(f"B{rr}", row_data['period'], font=font_total,
                  fill=fill_section, align=left)
        set_cell(f"C{rr}", None, font=font_value, fill=row_fill, fmt=INT_FMT)          # container vessels
        set_cell(f"D{rr}", None, font=font_value, fill=row_fill, fmt=NUM_FMT)          # imp teus
        set_cell(f"E{rr}", None, font=font_value, fill=row_fill, fmt=NUM_FMT)          # exp teus
        set_cell(f"F{rr}", f"=SUM(D{rr}:E{rr})", font=font_total, fill=fill_total, fmt=NUM_FMT)  # total teus
        set_cell(f"G{rr}", row_data['bulk_vessels'], font=font_value, fill=row_fill, fmt=INT_FMT)
        set_cell(f"H{rr}", row_data['dry_bulk_tons'], font=font_value, fill=row_fill, fmt=NUM_FMT)
        set_cell(f"I{rr}", row_data['break_bulk_tons'], font=font_value, fill=row_fill, fmt=NUM_FMT)
        set_cell(f"J{rr}", row_data['liquid_bulk_tons'], font=font_value, fill=row_fill, fmt=NUM_FMT)
        set_cell(f"K{rr}", row_data['bulk_total_tons'], font=font_total, fill=fill_total, fmt=NUM_FMT)

    last_traffic_row = r2 + len(traffic) - 1

    nr = last_traffic_row + 2

    # ---- YARD INVENTORY IN TEUS ----
    # Horizontal layout: IMPORT | EXPORT | TRANSHIPMENT | TOTAL as column headers
    set_cell(f"A{nr}", "YARD INVENTORY IN TEUS", font=font_section, fill=fill_section, align=left)
    set_cell(f"B{nr}", "IMPORT",       font=font_header, fill=fill_header)
    set_cell(f"C{nr}", "EXPORT",       font=font_header, fill=fill_header)
    set_cell(f"D{nr}", "TRANSHIPMENT", font=font_header, fill=fill_header)
    set_cell(f"E{nr}", "TOTAL",        font=font_header, fill=fill_total)
    set_cell(f"A{nr+1}", None, fill=fill_section)
    set_cell(f"B{nr+1}", None, font=font_value, fill=fill_white, fmt=NUM_FMT)
    set_cell(f"C{nr+1}", None, font=font_value, fill=fill_white, fmt=NUM_FMT)
    set_cell(f"D{nr+1}", None, font=font_value, fill=fill_white, fmt=NUM_FMT)
    set_cell(f"E{nr+1}", f"=SUM(B{nr+1}:D{nr+1})", font=font_total, fill=fill_total, fmt=NUM_FMT)
    nr = nr + 2 + 1

    # ---- GATE MOVEMENTS ----
    merge(f"A{nr}:A{nr+2}", "GATE MOVEMENTS", font=font_section, fill=fill_section, align=left)
    merge(f"B{nr}:C{nr}", "GATE MOVEMENT IN TEUs", font=font_header, fill=fill_header)
    merge(f"D{nr}:F{nr}", "NO. OF TRUCKS",         font=font_header, fill=fill_header)
    set_cell(f"B{nr+1}", "IN",      font=font_header, fill=fill_header)
    set_cell(f"C{nr+1}", "OUT",     font=font_header, fill=fill_header)
    set_cell(f"D{nr+1}", "IN",      font=font_header, fill=fill_header)
    set_cell(f"E{nr+1}", "OUT",     font=font_header, fill=fill_header)
    set_cell(f"F{nr+1}", "AVG TAT", font=font_header, fill=fill_total)
    for col in ("B", "C", "D", "E", "F"):
        set_cell(f"{col}{nr+2}", None, font=font_value, fill=fill_white, fmt=NUM_FMT)
    nr = nr + 3 + 1

    # ---- ICD PENDENCY ----
    icd_labels = ["TKD", "OTHERS", "TOTAL"]
    merge(f"A{nr}:A{nr+len(icd_labels)}", "ICD PENDENCY", font=font_section, fill=fill_section, align=left)
    set_cell(f"B{nr}", "DESTINATION", font=font_header, fill=fill_header, align=left)
    set_cell(f"C{nr}", "TEUS",        font=font_header, fill=fill_header)
    for i, lbl in enumerate(icd_labels):
        rr = nr + 1 + i
        is_total = lbl == "TOTAL"
        if is_total:
            set_cell(f"B{rr}", lbl, font=font_total, fill=fill_total, align=left)
            set_cell(f"C{rr}", f"=SUM(C{nr+1}:C{nr+2})", font=font_total, fill=fill_total, fmt=NUM_FMT)
        else:
            set_cell(f"B{rr}", lbl, font=font_normal, fill=fill_white, align=left)
            set_cell(f"C{rr}", None, font=font_value, fill=fill_white, fmt=NUM_FMT)
    nr = nr + len(icd_labels) + 2

    # ---- CFS PENDENCY ----
    merge(f"A{nr}:A{nr+1}", "CFS PENDENCY", font=font_section, fill=fill_section, align=left)
    set_cell(f"B{nr}", "DESTINATION", font=font_header, fill=fill_header, align=left)
    set_cell(f"C{nr}", "TEUS",        font=font_header, fill=fill_header)
    set_cell(f"B{nr+1}", "OTHERS", font=font_normal, fill=fill_white, align=left)
    set_cell(f"C{nr+1}", None, font=font_value, fill=fill_white, fmt=NUM_FMT)
    nr = nr + 2 + 1

    # ---- IMPORT EMPTY ----
    merge(f"A{nr}:A{nr+1}", "IMPORT EMPTY", font=font_section, fill=fill_section, align=left)
    set_cell(f"B{nr}", "CATEGORY", font=font_header, fill=fill_header, align=left)
    set_cell(f"C{nr}", "TEUS",     font=font_header, fill=fill_header)
    set_cell(f"B{nr+1}", "INYARD", font=font_normal, fill=fill_white, align=left)
    set_cell(f"C{nr+1}", None, font=font_value, fill=fill_white, fmt=NUM_FMT)
    nr = nr + 2 + 1

    # ---- REEFER SLOTS ----
    merge(f"A{nr}:A{nr+1}", "REEFER SLOTS", font=font_section, fill=fill_section, align=left)
    set_cell(f"B{nr}", "TOTAL",     font=font_header, fill=fill_header)
    set_cell(f"C{nr}", "OCCUPIED",  font=font_header, fill=fill_header)
    set_cell(f"D{nr}", "AVAILABLE", font=font_header, fill=fill_total)
    set_cell(f"B{nr+1}", None, font=font_value, fill=fill_white, fmt=INT_FMT)
    set_cell(f"C{nr+1}", None, font=font_value, fill=fill_white, fmt=INT_FMT)
    set_cell(f"D{nr+1}", f"=B{nr+1}-C{nr+1}", font=font_total, fill=fill_total, fmt=INT_FMT)
    nr = nr + 2 + 1





    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    return wb


@bp.route('/api/module/RP01/jjltpl/export')
@login_required
def jjltpl_export():
    selected_date = _jjltpl_parse_date(request.args.get('date'))
    terminal = request.args.get('terminal', DEFAULT_TERMINAL)
    payload = _jjltpl_report_payload(selected_date, terminal)

    wb = _jjltpl_build_workbook(payload)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Terminal_Traffic_Report_{terminal}_{payload['date']}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )